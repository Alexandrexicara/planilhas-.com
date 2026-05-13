import sys
import os

# Configurar UTF-8 para suportar emojis no Windows (antes de outros imports)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import sqlite3
import zipfile
import shutil
import subprocess
import time
import importlib
import threading
import webbrowser
import runpy
import requests
from functools import wraps
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import json
from datetime import datetime
import socket
import hashlib
import base64

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, send_from_directory, session, g

print("=== APP INICIANDO ===")
print("Python version:", sys.version)
print("Working directory:", os.getcwd())
print("DATABASE_URL encontrado:", bool(os.environ.get('DATABASE_URL')))

try:
    from planilhas_paths import runtime_dir as _runtime_dir, ensure_from_resource as _ensure_from_resource, is_frozen as _is_frozen
    print("[OK] planilhas_paths importado")
except Exception as e:
    print("[ERRO] Erro ao importar planilhas_paths:", e)
    raise

print("=== IMPORTS BÁSICOS OK ===")

# SQLite temporariamente para testar
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///acesso_web.db'
print("DATABASE_URL encontrado: False")
# Usa PostgreSQL se DATABASE_URL existir (Render), senão SQLite local (fallback)
from web_access_db_postgres import (
    init_db as _init_access_db,
    ensure_superadmin as _ensure_superadmin,
    authenticate as _auth_user,
    get_user as _get_user,
    get_organization as _get_org,
    organization_has_access as _org_has_access,
    create_organization as _create_org,
    create_user as _create_user,
    create_invite as _create_invite,
    redeem_invite as _redeem_invite,
)

print("=== IMPORTS DE BANCO OK ===")

def gerar_url_curta(base_url, codigo_convite):
    """Gera URL curta com preview usando hash base64"""
    # Criar hash único para o convite
    hash_obj = hashlib.md5(f"{codigo_convite}_{time.time()}".encode())
    short_code = base64.urlsafe_b64encode(hash_obj.digest()[:6]).decode('utf-8').rstrip('=')
    
    # URL curta simulada (em produção poderia usar bit.ly, tinyurl, etc)
    url_curta = f"https://inv.pt/{short_code}"
    
    return {
        'curta': url_curta,
        'preview': f"Convite para organização - Clique para aceitar",
        'qr_code': f"data:image/png;base64,{gerar_qr_code_base64(base_url)}" if gerar_qr_code_base64 else None
    }

def gerar_qr_code_base64(url):
    """Gera QR Code em base64 (requiere qrcode pillow)"""
    try:
        import qrcode
        from io import BytesIO
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()
        return img_str
    except ImportError:
        return None
    except Exception as e:
        print(f"Erro ao gerar QR code: {e}")
        return None

# Verificar DATABASE_URL
DATABASE_URL = os.environ.get('DATABASE_URL')
print(f"DATABASE_URL encontrado: {DATABASE_URL is not None}")

if not DATABASE_URL:
    print("[AVISO] DATABASE_URL não configurado, usando SQLite fallback")

from pagbank_client import client_from_env as _pagbank_from_env

# Importacao segura dos modulos desktop (podem falhar em ambiente server/headless)
MODULOS_IMPORTACAO = {}
ERROS_IMPORTACAO = {}


def importar_modulo(nome_modulo):
    try:
        modulo = importlib.import_module(nome_modulo)
        MODULOS_IMPORTACAO[nome_modulo] = True
        return modulo
    except Exception as e:
        MODULOS_IMPORTACAO[nome_modulo] = False
        ERROS_IMPORTACAO[nome_modulo] = str(e)
        print(f"[AVISO] Nao foi possivel importar {nome_modulo}: {e}")
        return None


sistema = importar_modulo('sistema')
sistema_plus = importar_modulo('sistema_plus')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IS_VERCEL = bool(os.environ.get('VERCEL'))
IS_RENDER = bool(os.environ.get('RENDER'))
RUNTIME_DIR = os.path.join('/tmp', 'planilhas') if IS_VERCEL else _runtime_dir()
UPLOAD_DIR = os.path.join(RUNTIME_DIR, 'uploads')
TEMP_IMAGES_DIR = os.path.join(RUNTIME_DIR, 'temp_images')
STATIC_UPLOADS_DIR = os.path.join(RUNTIME_DIR, 'static_uploads')
BUNDLED_DB_PATH = os.path.join(BASE_DIR, 'banco_plus.db')
DB_PATH = os.path.join(RUNTIME_DIR, 'banco_plus.db') if IS_VERCEL else _ensure_from_resource('banco_plus.db')

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "sistema_plus_2024")
app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Banco de acesso (login/roles/pagamento) para a interface web
_init_access_db()
_ensure_superadmin(os.environ.get("PLANILHAS_CREATOR_EMAIL"), os.environ.get("PLANILHAS_CREATOR_PASSWORD"))

# Garantir admin padrão em qualquer ambiente (local, exe, Render)
try:
    _ensure_superadmin("admin@planilhas.com", "admin123")
    _ensure_superadmin("superadmin@planilhas.com", "GpA1XmI86lGB309W")
    print("[INIT] Admins padrão garantidos: admin@planilhas.com / superadmin@planilhas.com")
except Exception as _e:
    print(f"[AVISO] Não foi possível criar admins padrões: {_e}")

# Inicializar tabela de licenças desktop (anti-pirataria)
try:
    import desktop_license as _desktop_license
    _desktop_license.init_license_table()
    print("[INIT] Tabela desktop_licenses pronta")
except Exception as _e:
    print(f"[AVISO] Não foi possível inicializar desktop_licenses: {_e}")
    _desktop_license = None

# Pastas de runtime (serverless so pode escrever em /tmp)
for runtime_path in [RUNTIME_DIR, UPLOAD_DIR, TEMP_IMAGES_DIR, STATIC_UPLOADS_DIR]:
    try:
        os.makedirs(runtime_path, exist_ok=True)
    except Exception as e:
        print(f"[AVISO] Nao foi possivel criar pasta {runtime_path}: {e}")

# Em Vercel, copia o banco empacotado para /tmp na primeira execucao
if IS_VERCEL and os.path.exists(BUNDLED_DB_PATH) and not os.path.exists(DB_PATH):
    try:
        shutil.copy2(BUNDLED_DB_PATH, DB_PATH)
    except Exception as e:
        print(f"[AVISO] Nao foi possivel copiar banco para runtime: {e}")

# Configuracao do sistema
SISTEMA_CONFIG = {
    'nome': 'planilhas.com',
    'valor_original': 5000.00,
    'valor_promocional': 4500.00,
    'desconto': 10,
    'versao': '2.0',
    'recursos': [
        'Importacao automatica de Excel',
        'Extracao automatica de imagens',
        'Catalogo web completo',
        'Gestao de produtos',
        'Relatorios detalhados'
    ]
}

BUILD_INFO = {
    'render': IS_RENDER,
    'service': os.environ.get('RENDER_SERVICE_NAME', ''),
    'vercel': IS_VERCEL,
    'region': os.environ.get('VERCEL_REGION', ''),
    'commit': os.environ.get('RENDER_GIT_COMMIT', '') or os.environ.get('VERCEL_GIT_COMMIT_SHA', ''),
}

# Inicialização no Render (executado pelo Gunicorn)
if IS_RENDER:
    print("=== RENDER DETECTADO - INICIALIZANDO BANCO ===")
    try:
        _init_access_db()
        print("Banco PostgreSQL inicializado")
        
        # Criar superadmin se não existir
        result = _ensure_superadmin("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        print(f"Superadmin garantido: {result}")
        
        # Criar admin padrão (admin@planilhas.com / admin123)
        result_admin = _ensure_superadmin("admin@planilhas.com", "admin123")
        print(f"Admin garantido: {result_admin}")
        
        # Testar autenticação
        test_user = _auth_user("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        if test_user:
            print("Superadmin autenticado com sucesso!")
        else:
            print("ERRO: Superadmin não autenticou!")
        
        test_admin = _auth_user("admin@planilhas.com", "admin123")
        if test_admin:
            print("Admin autenticado com sucesso!")
        else:
            print("ERRO: Admin não autenticou!")
            
    except Exception as e:
        print(f"ERRO NA INICIALIZACAO DO RENDER: {e}")
        import traceback
        traceback.print_exc()

DESKTOP_RELEASES_DIRS = [
    os.path.join(BASE_DIR, 'releases'),
    os.path.join(BASE_DIR, 'dist'),
]
DESKTOP_POLICY = {
    'max_maquinas': 10,
    'atualizacao_requer_compra': True
}


def get_user_data_dir():
    """Diretorio gravavel pelo usuario (evita Program Files)."""
    base = os.environ.get('LOCALAPPDATA') or os.environ.get('APPDATA') or os.path.expanduser('~')
    return os.path.join(base, 'planilhas.com')


def log_desktop(msg):
    """Log simples para diagnostico quando o exe roda sem console."""
    try:
        log_dir = os.path.join(get_user_data_dir(), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        log_path = os.path.join(log_dir, 'planilhas_desktop.log')
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


def abrir_navegador_quando_pronto(url, host, port, timeout_segundos=12):
    """Abre o navegador quando a porta estiver aceitando conexoes."""
    def _abrir():
        deadline = time.time() + timeout_segundos
        while time.time() < deadline:
            try:
                with socket.create_connection((host, port), timeout=0.5):
                    break
            except Exception:
                time.sleep(0.25)

        try:
            webbrowser.open(url, new=1)
            log_desktop(f"Navegador aberto: {url}")
        except Exception as e:
            log_desktop(f"[AVISO] Nao foi possivel abrir navegador automaticamente: {e}")

    threading.Thread(target=_abrir, daemon=True).start()

def get_db_connection():
    """Conexao com o banco de dados (com isolamento multi-tenant)"""
    # Garantir que o banco existe no Render
    if IS_RENDER and not os.path.exists(DB_PATH):
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        conn = sqlite3.connect(DB_PATH)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS produtos_plus (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                organization_id INTEGER,
                nome TEXT,
                descricao TEXT,
                preco REAL,
                imagem TEXT,
                cliente TEXT,
                data_importacao TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        conn.close()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Migração automática: adicionar organization_id se a coluna não existir
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(produtos_plus)").fetchall()]
        if cols and 'organization_id' not in cols:
            print("[MIGRAÇÃO] Adicionando coluna organization_id em produtos_plus")
            conn.execute("ALTER TABLE produtos_plus ADD COLUMN organization_id INTEGER")
            conn.commit()
    except Exception as _e:
        print(f"[AVISO] Migração organization_id falhou: {_e}")

    return conn


def _user_org_id():
    """Retorna o organization_id do usuário logado, ou None."""
    user = _current_user()
    if not user:
        return None
    return user.get("organization_id")


def _filtro_org_sql(alias=""):
    """Devolve (clausula_sql, params) para isolar dados pela organização do user.
    Superadmin vê tudo. Usuários sem organização veem somente itens sem org.
    """
    user = _current_user()
    pref = (alias + ".") if alias else ""
    if _is_superadmin(user):
        return "", []
    org_id = _user_org_id()
    if org_id is None:
        return f"{pref}organization_id IS NULL", []
    return f"({pref}organization_id = ? OR {pref}organization_id IS NULL)", [org_id]


def _current_user():
    return getattr(g, "current_user", None)


def _is_superadmin(user):
    return bool(user) and user.get("role") == "superadmin"

def _is_admin_or_superadmin(user):
    return bool(user) and user.get("role") in ["superadmin", "admin"]


def _login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = _current_user()
        if not user:
            nxt = request.full_path if request.query_string else request.path
            return redirect(url_for("comecar", next=nxt))
        return view(*args, **kwargs)

    return wrapped


def _paid_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = _current_user()
        if not user:
            nxt = request.full_path if request.query_string else request.path
            return redirect(url_for("comecar", next=nxt))
        if _is_admin_or_superadmin(user):
            return view(*args, **kwargs)
        org_id = user.get("organization_id")
        if org_id and _org_has_access(org_id):
            return view(*args, **kwargs)
        nxt = request.full_path if request.query_string else request.path
        return redirect(url_for("pagamento", next=nxt))

    return wrapped


def _role_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = _current_user()
            if not user:
                nxt = request.full_path if request.query_string else request.path
                return redirect(url_for("comecar", next=nxt))
            if _is_admin_or_superadmin(user) or user.get("role") in roles:
                return view(*args, **kwargs)
            return ("Acesso negado", 403)

        return wrapped

    return decorator


@app.before_request
def _load_current_user():
    g.current_user = None
    user_id = session.get("user_id")
    if user_id:
        user = _get_user(user_id)
        if user and int(user.get("ativo", 1)) == 1:
            g.current_user = user


def _build_system_status():
    dados_sistema = {
        "status": "sucesso",
        "mensagem": "Sistema pronto.",
        "timestamp": datetime.now().isoformat(),
        "modulos_carregados": [],
        "build_info": BUILD_INFO,
        "distribuicao_desktop": get_desktop_distribution_info(),
    }

    try:
        if sistema is not None:
            dados_sistema["modulos_carregados"].append("[OK] sistema.py carregado")
        else:
            erro = ERROS_IMPORTACAO.get("sistema", "modulo nao carregado")
            dados_sistema["modulos_carregados"].append(f"[ERRO] sistema.py: {erro[:80]}")
    except Exception as e:
        dados_sistema["modulos_carregados"].append(f"[ERRO] sistema.py: {str(e)[:50]}")

    try:
        if sistema_plus is not None and hasattr(sistema_plus, "contar_produtos_plus"):
            dados_sistema["total_produtos"] = sistema_plus.contar_produtos_plus()
        if sistema_plus is not None and hasattr(sistema_plus, "contar_planilhas_plus"):
            dados_sistema["total_planilhas"] = sistema_plus.contar_planilhas_plus()

        if sistema_plus is not None:
            dados_sistema["modulos_carregados"].append("[OK] sistema_plus.py carregado")
        else:
            erro = ERROS_IMPORTACAO.get("sistema_plus", "modulo nao carregado")
            dados_sistema["modulos_carregados"].append(f"[ERRO] sistema_plus.py: {erro[:80]}")
    except Exception as e:
        dados_sistema["modulos_carregados"].append(f"[ERRO] sistema_plus.py: {str(e)[:50]}")

    return dados_sistema


@app.route("/comecar")
def comecar():
    """Tela de login."""
    nxt = request.args.get("next")
    if nxt:
        session["next"] = nxt
    return render_template("login_separado.html", config=SISTEMA_CONFIG, next=nxt or "")


@app.route("/cadastro")
def cadastro():
    """Tela de registro."""
    nxt = request.args.get("next")
    codigo_convite = request.args.get("convite")
    
    if nxt:
        session["next"] = nxt
    
    # Se há código de convite, validar e mostrar informações
    convite_info = None
    if codigo_convite:
        try:
            conn = _get_db_connection()
            convite = conn.execute(
                """
                SELECT i.code, i.role, i.email, i.created_at, o.nome as org_nome
                FROM invites i
                JOIN organizations o ON i.organization_id = o.id
                WHERE i.code = ? AND i.used_at IS NULL
                """,
                (codigo_convite,)
            ).fetchone()
            
            if convite:
                convite_info = {
                    'code': convite['code'],
                    'role': convite['role'],
                    'email': convite['email'],
                    'org_nome': convite['org_nome'],
                    'created_at': convite['created_at']
                }
        except Exception as e:
            print(f"Erro ao validar convite: {e}")
    
    return render_template(
        "registro_separado.html", 
        config=SISTEMA_CONFIG, 
        next=nxt or "",
        convite=convite_info
    )


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    return redirect(url_for("index"))


@app.route("/create-superadmin-manual")
def create_superadmin_manual():
    """Criar superadmin imediatamente sem login"""
    try:
        print("=== CRIANDO SUPERADMIN MANUALMENTE ===")
        
        # Inicializar banco primeiro
        _init_access_db()
        print("Banco inicializado")
        
        # Criar superadmin
        result = _ensure_superadmin("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        print(f"Resultado superadmin: {result}")
        
        # Criar usuário comum
        try:
            user_result = _create_user(
                email="santossilvac990@gmail.com",
                senha="celio48santos",
                nome="Usuario Teste",
                role="user"
            )
            print(f"Resultado usuário: {user_result}")
        except Exception as e:
            print(f"Usuário já existe ou erro: {e}")
        
        # Testar autenticação superadmin
        test_superadmin = _auth_user("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        if test_superadmin:
            print("Superadmin criado e autenticado com sucesso!")
        else:
            print("ERRO: Superadmin não autenticou!")
            
        # Testar autenticação usuário
        test_user = _auth_user("santossilvac990@gmail.com", "celio48santos")
        if test_user:
            print("Usuário criado e autenticado com sucesso!")
        else:
            print("ERRO: Usuário não autenticou!")
        
        return "Superadmin e usuário criados com sucesso! Tente fazer login agora."
        
    except Exception as e:
        print(f"ERRO AO CRIAR SUPERADMIN: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro: {e}", 500


@app.route("/admin/create-superadmin", methods=["GET", "POST"])
def create_superadmin():
    """Endpoint para criar superadmin manualmente (debug)"""
    if request.method == "GET":
        return render_template("create_superadmin.html")
    
    try:
        print("=== CRIANDO SUPERADMIN MANUALMENTE ===")
        
        # Criar superadmin
        result = _ensure_superadmin("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        print(f"Resultado: {result}")
        
        # Testar autenticação
        test_user = _auth_user("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        if test_user:
            print("Superadmin criado e autenticado com sucesso!")
        else:
            print("ERRO: Superadmin criado mas não autentica!")
        
        return redirect(url_for("index"))
        
    except Exception as e:
        print(f"ERRO AO CRIAR SUPERADMIN: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro: {e}", 500


@app.route("/login", methods=["GET"])
def login_page():
    """Página de login"""
    print("=== DEBUG LOGIN PAGE (TEMPLATE) ===")
    print("Renderizando template: login.html")
    return render_template('login.html')

@app.route("/cadastro")
def cadastro_page():
    """Página de cadastro"""
    print("=== DEBUG CADASTRO PAGE (TEMPLATE) ===")
    print("Renderizando template: cadastro.html")
    return render_template('cadastro.html')

@app.route("/login", methods=["POST"])
def login():
    print(f"=== DEBUG ROTA /LOGIN ===")
    print(f"Form data completo: {dict(request.form)}")
    print(f"Request method: {request.method}")
    
    try:
        email = (request.form.get("email") or "").strip()
        senha = request.form.get("senha") or ""
        nxt = (request.form.get("next") or session.pop("next", "") or "").strip()

        print(f"Email extraído: '{email}'")
        print(f"Senha extraída: '{senha}'")
        print(f"Next: '{nxt}'")

        print("=== DEBUG COMPLETO LOGIN ===")
        print(f"EMAIL RECEBIDO: '{email}'")
        print(f"SENHA RECEBIDA: '{senha[:3]}***' (len={len(senha)})")
        
        print("Chamando _auth_user...")
        user = _auth_user(email, senha)
        print(f"Resultado _auth_user: {user}")
        print(f"Tipo do resultado: {type(user)}")
        
        if user is None:
            print("=== USUÁRIO É NONE ===")
            print("VERIFICANDO SE USUÁRIO EXISTE NO BANCO...")
            
            # Verificar diretamente no banco
            try:
                from web_access_db_postgres import connect
                conn = connect()
                db_url = os.environ.get('DATABASE_URL')
                is_postgres = bool(db_url)
                
                if is_postgres:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute("SELECT id, email, role FROM users WHERE email = %s", (email.strip().lower(),))
                        row = cur.fetchone()
                        print(f"POSTGRES - Usuário encontrado: {row}")
                else:
                    row = conn.execute("SELECT id, email, role FROM users WHERE email = ?", (email.strip().lower(),)).fetchone()
                    print(f"SQLITE - Usuário encontrado: {row}")
                
                conn.close()
                
                # Listar todos os usuários
                conn2 = connect()
                if is_postgres:
                    with conn2.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute("SELECT id, email, role FROM users ORDER BY id")
                        all_users = cur.fetchall()
                        print(f"POSTGRES - Todos usuários: {all_users}")
                else:
                    all_users = conn2.execute("SELECT id, email, role FROM users ORDER BY id").fetchall()
                    print(f"SQLITE - Todos usuários: {all_users}")
                conn2.close()
                
            except Exception as e:
                print(f"ERRO AO VERIFICAR BANCO: {e}")
            
            print("Usuário não encontrado ou senha inválida")
            return jsonify({"success": False, "message": "Email ou senha inválidos"})

        session["user_id"] = user["id"]
        print(f"Session user_id definido: {user['id']}")

        if _is_admin_or_superadmin(user):
            print("Usuário é admin/superadmin")
            return jsonify({"success": True, "message": "Login realizado", "redirect": nxt or url_for("sistema_dashboard")})

        org_id = user.get("organization_id")
        if not org_id:
            print("Usuário sem organização")
            return jsonify({"success": True, "message": "Login realizado", "redirect": nxt or url_for("sistema_dashboard")})

        if _org_has_access(org_id):
            print("Organização tem acesso")
            return jsonify({"success": True, "message": "Login realizado", "redirect": nxt or url_for("sistema_dashboard")})

        print("Organização sem acesso - pagamento pendente")
        return jsonify({"success": True, "message": "Login realizado. Pagamento pendente.", "redirect": url_for("pagamento")})
    
    except Exception as e:
        print(f"ERRO NO LOGIN: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Erro no servidor: {str(e)}"}), 500


@app.route("/registro", methods=["POST"])
def registro():
    try:
        print(f"=== DEBUG REGISTRO ===")
        print(f"Form data: {dict(request.form)}")
        
        nome = (request.form.get("nome") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        senha = request.form.get("senha") or ""
        senha2 = request.form.get("senha2") or request.form.get("confirmar_senha") or ""
        codigo_convite = (request.form.get("codigo_convite") or "").strip()
        org_nome = (request.form.get("org_nome") or "").strip()
        cpf = (request.form.get("cpf") or "").strip()
        nxt = (request.form.get("next") or session.pop("next", "") or "").strip()

        print(f"Dados extraídos: nome={nome}, email={email}, org_nome={org_nome}, senha_len={len(senha)}")

        if not nome or not email or not senha:
            return jsonify({"success": False, "message": "Preencha nome, email e senha"})
        if senha2 and senha2 != senha:
            return jsonify({"success": False, "message": "As senhas não conferem"})

        if codigo_convite:
            print(f"Tentando registrar com convite: {codigo_convite}")
            ok, msg, user_id = _redeem_invite(codigo_convite, nome, email, senha)
            if not ok:
                return jsonify({"success": False, "message": msg})
            session["user_id"] = user_id
            return jsonify({"success": True, "message": msg, "redirect": nxt or url_for("sistema_dashboard")})

        if not org_nome:
            return jsonify({"success": False, "message": "Informe o nome da empresa (ou use um código de convite)"})

        print(f"Criando organização: {org_nome}")
        org_id = _create_org(org_nome, payment_amount=SISTEMA_CONFIG.get("valor_promocional", 50.00))
        print(f"Organização criada com ID: {org_id}")
        
        print(f"Criando usuário: {nome}, {email}")
        user_id = _create_user(org_id, nome, email, senha, role="owner")
        print(f"Usuário criado com ID: {user_id}")
        
        session["user_id"] = user_id
        print(f"Session user_id definido: {user_id}")

        # Tentar gerar cobrança Pix automaticamente (se PagBank estiver configurado)
        try:
            pagbank = _pagbank_from_env()
            pix_key = os.environ.get("PAGBANK_PIX_KEY") or os.environ.get("PAGBANK_RECEIVER_PIX_KEY") or ""
            if pagbank.is_configured():
                # Tenta criar checkout completo primeiro
                charge = pagbank.create_checkout_charge(
                    amount=SISTEMA_CONFIG.get("valor_promocional", 50.00),
                    payer_name=nome,
                    payer_email=email,
                    payer_cpf=cpf or "12345678909",
                    description=f"planilhas.com - Licença vitalícia ({org_nome})",
                    redirect_url=url_for("pagamento", _external=True),
                    webhook_url=url_for("webhook_pagbank", _external=True),
                )
                
                if charge.get("charge_id"):
                    _org_set_pending(org_id, charge.get("charge_id"), charge.get("qr_code_base64"), charge.get("pix_key"))
                elif charge.get("txid"):
                    _org_set_pending(org_id, charge.get("txid"), charge.get("qr_code_base64"), charge.get("pix_key"))
        except Exception:
            pass

        print(f"Cadastro finalizado com sucesso! Redirecionando para pagamento...")
        return jsonify(
            {"success": True, "message": "Cadastro realizado. Finalize o pagamento para liberar o acesso.", "redirect": url_for("pagamento")}
        )
    except sqlite3.IntegrityError:
        print(f"ERRO: Email já cadastrado: {email}")
        return jsonify({"success": False, "message": "Este email já está cadastrado"})
    except Exception as e:
        print(f"ERRO NO REGISTRO: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Erro ao registrar: {str(e)}"})


@app.route("/pagamento")
@_login_required
def pagamento():
    user = _current_user()
    if _is_admin_or_superadmin(user):
        return redirect(url_for("sistema_dashboard"))

    org = _get_org(user.get("organization_id"))
    if not org:
        return ("Organização não encontrada", 400)

    if org.get("payment_status") == "paid":
        return redirect(url_for("sistema_dashboard"))

    qr_data_uri = None
    if org.get("payment_qr_base64"):
        qr_data_uri = f"data:image/png;base64,{org['payment_qr_base64']}"

    return render_template(
        "pagamento.html",
        config=SISTEMA_CONFIG,
        org=org,
        qr_data_uri=qr_data_uri,
        pix_key=org.get("payment_pix_key") or (os.environ.get("PAGBANK_PIX_KEY") or ""),
        next=request.args.get("next", ""),
    )


@app.route("/api/pagamento/gerar", methods=["POST"])
@_login_required
@_role_required("owner", "admin")
def api_pagamento_gerar():
    user = _current_user()
    org = _get_org(user.get("organization_id"))
    if not org:
        return jsonify({"success": False, "message": "Organização não encontrada"}), 400

    if org.get("payment_status") == "paid":
        return jsonify({"success": True, "message": "Pagamento já liberado", "status": "paid"})

    pagbank = _pagbank_from_env()
    pix_key = os.environ.get("PAGBANK_PIX_KEY") or os.environ.get("PAGBANK_RECEIVER_PIX_KEY") or ""
    if not pagbank.is_configured() or not pix_key:
        return jsonify({"success": False, "message": "PagBank não configurado no servidor"}), 400

    try:
        charge = pagbank.create_checkout_charge(
            amount=org.get("payment_amount") or SISTEMA_CONFIG.get("valor_promocional", 50.00),
            payer_name=user.get("nome"),
            payer_email=user.get("email"),
            payer_cpf="12345678909",
            description=f"planilhas.com - Licença vitalícia ({org.get('nome')})",
            redirect_url=url_for("pagamento", _external=True),
            webhook_url=url_for("webhook_pagbank", _external=True),
        )
        
        if charge.get("charge_id"):
            _org_set_pending(org["id"], charge.get("charge_id"), charge.get("qr_code_base64"), charge.get("pix_key"))
            return jsonify(
                {
                    "success": True,
                    "message": "Pagamento gerado com sucesso",
                    "charge_id": charge.get("charge_id"),
                    "checkout_url": charge.get("checkout_url"),
                    "payment_urls": charge.get("payment_urls", []),
                    "links": charge.get("links", []),
                    "status": charge.get("status", "pending"),
                }
            )
        elif charge.get("txid"):
            _org_set_pending(org["id"], charge.get("txid"), charge.get("qr_code_base64"), charge.get("pix_key"))
            return jsonify(
                {
                    "success": True,
                    "message": "Cobrança PIX gerada",
                    "txid": charge.get("txid"),
                    "qr_data_uri": f"data:image/png;base64,{charge.get('qr_code_base64')}" if charge.get('qr_code_base64') else None,
                    "pix_key": charge.get("pix_key"),
                    "status": "pending",
                }
            )
        else:
            return jsonify({"success": False, "message": "Não foi possível gerar a cobrança"}), 400
    except Exception as e:
        return jsonify({"success": False, "message": f"Erro ao gerar cobrança: {str(e)}"}), 400


@app.route("/webhook/pagbank", methods=["POST"])
def webhook_pagbank():
    """Webhook para receber notificações do PagBank"""
    try:
        data = request.get_json()
        if not data:
            return "OK", 200
        
        # Log para debug
        print(f"Webhook PagBank recebido: {data}")
        
        # Verificar se é notificação de pagamento
        charge_id = data.get("charge_id") or data.get("id")
        status = data.get("status", "").upper()
        
        if charge_id and status in ["PAID", "APPROVED", "COMPLETED"]:
            # Atualizar status no banco
            conn = get_db_connection()
            try:
                conn.execute(
                    "UPDATE organizations SET payment_status = 'paid', payment_updated_at = ? WHERE payment_txid = ?",
                    (_utcnow_str(), charge_id)
                )
                conn.commit()
                print(f"Pagamento confirmado: {charge_id}")
            finally:
                conn.close()
        
        return "OK", 200
    except Exception as e:
        print(f"Erro no webhook PagBank: {e}")
        return "OK", 200  # Sempre retornar 200 para não reenviar


@app.route("/api/user/check")
def api_user_check():
    """API para verificar se usuário está autenticado"""
    user = _current_user()
    return jsonify({
        "authenticated": bool(user),
        "email": user.get("email") if user else None,
        "role": user.get("role") if user else None
    })

@app.route("/api/pagamento/status")
@_login_required
def api_pagamento_status():
    user = _current_user()
    if _is_admin_or_superadmin(user):
        return jsonify({"success": True, "status": "paid"})

    org = _get_org(user.get("organization_id"))
    if not org:
        return jsonify({"success": False, "message": "Organização não encontrada"}), 400

    if org.get("payment_status") == "paid":
        return jsonify({"success": True, "status": "paid"})

    txid = org.get("payment_txid")
    if not txid:
        return jsonify({"success": True, "status": org.get("payment_status") or "unpaid"})

    pagbank = _pagbank_from_env()
    if not pagbank.is_configured():
        return jsonify({"success": True, "status": org.get("payment_status") or "pending"})

    try:
        status = pagbank.get_charge_status(txid)
        if status == "CONCLUIDA":
            _org_set_paid(org["id"])
            return jsonify({"success": True, "status": "paid"})
        if status == "ATIVA":
            return jsonify({"success": True, "status": "pending"})
        return jsonify({"success": True, "status": "pending", "pagbank_status": status})
    except Exception:
        return jsonify({"success": True, "status": org.get("payment_status") or "pending"})




@app.route("/admin/colaboradores")
@_login_required
@_paid_required
@_role_required("owner", "admin")
def admin_colaboradores():
    user = _current_user()
    org = _get_org(user.get("organization_id"))
    return render_template(
        "admin_colaboradores.html",
        config=SISTEMA_CONFIG,
        org=org,
        invites=_list_invites(org["id"]) if org else [],
        users=_list_users(org["id"]) if org else [],
    )


@app.route("/i/<short_code>")
def redirect_convite_curto(short_code):
    """Redireciona URL curta para o convite completo"""
    try:
        # Buscar convite pelo código curto (em produção poderia ter tabela de mapeamento)
        conn = _get_db_connection()
        convites = conn.execute("SELECT code FROM invites WHERE used_at IS NULL ORDER BY created_at DESC LIMIT 100").fetchall()
        
        # Tentar encontrar convite correspondente (simplificado)
        for convite in convites:
            hash_obj = hashlib.md5(f"{convite['code']}_{int(time.time())}".encode())
            generated_short = base64.urlsafe_b64encode(hash_obj.digest()[:6]).decode('utf-8').rstrip('=')
            
            if generated_short == short_code:
                base_url = request.url_root.rstrip('/')
                return redirect(f"{base_url}/cadastro?convite={convite['code']}")
        
        return "Link inválido ou expirado", 404
    except Exception as e:
        print(f"Erro ao redirecionar link curto: {e}")
        return "Erro ao processar link", 500


@app.route("/admin/convites", methods=["POST"])
@_login_required
@_paid_required
@_role_required("owner", "admin")
def admin_criar_convite():
    user = _current_user()
    org_id = user.get("organization_id")
    if not org_id:
        return jsonify({"success": False, "message": "Organização não encontrada"}), 400
    role = (request.form.get("role") or "collab").strip().lower()
    email = (request.form.get("email") or "").strip()
    if role not in ("admin", "collab"):
        role = "collab"
    code = _create_invite(org_id, role=role, email=email or None)
    
    # Gerar link completo
    base_url = request.url_root.rstrip('/')
    invite_link = f"{base_url}/cadastro?convite={code}"
    
    # Gerar URL curta com preview
    url_info = gerar_url_curta(invite_link, code)
    
    return jsonify({
        "success": True, 
        "code": code,
        "link": invite_link,
        "link_curto": url_info['curta'],
        "preview": url_info['preview'],
        "qr_code": url_info['qr_code']
    })


def get_latest_desktop_exe():
    """Retorna o executavel de distribuicao em `releases/`.

    Prioriza nomes oficiais para evitar baixar builds antigos como `app.exe`.
    """
    preferred_names = ['Planilhas.exe', 'SistemaPlanilhas.exe']
    for base_dir in DESKTOP_RELEASES_DIRS:
        if not os.path.isdir(base_dir):
            continue
        for preferred_name in preferred_names:
            preferred_path = os.path.join(base_dir, preferred_name)
            if os.path.isfile(preferred_path):
                return {
                    'path': preferred_path,
                    'filename': preferred_name,
                    'updated_at': datetime.fromtimestamp(os.path.getmtime(preferred_path)).strftime('%Y-%m-%d %H:%M:%S')
                }

    exes = []
    for base_dir in DESKTOP_RELEASES_DIRS:
        if not os.path.isdir(base_dir):
            continue
        for entry in os.scandir(base_dir):
            if entry.is_file() and entry.name.lower().endswith('.exe'):
                exes.append(entry.path)

    if not exes:
        return None

    latest_path = max(exes, key=os.path.getmtime)
    return {
        'path': latest_path,
        'filename': os.path.basename(latest_path),
        'updated_at': datetime.fromtimestamp(os.path.getmtime(latest_path)).strftime('%Y-%m-%d %H:%M:%S')
    }


def get_desktop_distribution_info():
    """Dados para exibir download .exe e limite de maquinas."""
    exe = get_latest_desktop_exe()
    return {
        'disponivel': bool(exe),
        'arquivo': exe['filename'] if exe else None,
        'atualizado_em': exe['updated_at'] if exe else None,
        'download_url': '/download-exe' if exe else None,
        'max_maquinas': DESKTOP_POLICY['max_maquinas'],
        'atualizacao_requer_compra': DESKTOP_POLICY['atualizacao_requer_compra']
    }

def extract_images_from_excel(excel_path, output_dir):
    """Extrai imagens do arquivo Excel automaticamente"""
    images = []
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            for file in z.namelist():
                if file.startswith("xl/media/"):
                    name = os.path.basename(file)
                    path = os.path.join(output_dir, name)
                    
                    with open(path, "wb") as f:
                        f.write(z.read(file))
                    
                    images.append(path)
    except Exception as e:
        print(f"Erro ao extrair imagens: {e}")
    
    return images

def read_excel_with_images(excel_path):
    """Le Excel e associa imagens automaticamente"""
    temp_dir = TEMP_IMAGES_DIR
    os.makedirs(temp_dir, exist_ok=True)
    
    # 1. Extrair imagens
    images = extract_images_from_excel(excel_path, temp_dir)
    
    # 2. Ler dados do Excel
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    rows = []
    headers = []
    
    # Pegar cabecalhos
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Ler linhas de dados
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = str(value) if value is not None else ""
        
        # 3. Associar imagem por ordem
        img_index = row_idx - 2  # Ajuste para indice zero
        if img_index < len(images):
            row_data['picture'] = images[img_index]
        else:
            row_data['picture'] = None
        
        rows.append(row_data)
    
    return rows, images

def save_image_to_static(image_path):
    """Move imagem para pasta static e retorna URL"""
    if not image_path or not os.path.exists(image_path):
        return None
    
    filename = os.path.basename(image_path)
    new_path = os.path.join(STATIC_UPLOADS_DIR, filename)
    
    shutil.copy(image_path, new_path)
    return new_path if IS_VERCEL else f"/static/uploads/{filename}"

def import_products_to_db(products_data, organization_id=None):
    """Importa produtos para o banco de dados (vinculados à organização)."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    imported_count = 0
    
    for product in products_data:
        try:
            # Salvar imagem se existir
            picture_url = None
            if product.get('picture'):
                picture_url = save_image_to_static(product['picture'])
            
            # Inserir no banco
            cursor.execute("""
                INSERT INTO produtos_plus (
                    organization_id, cliente, arquivo_origem, codigo, descricao, peso, 
                    picture, data_importacao
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                organization_id,
                product.get('cliente', 'Web'),
                'upload_web',
                product.get('codigo', ''),
                product.get('descricao', ''),
                product.get('peso', ''),
                picture_url,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ))
            
            imported_count += 1
            
        except Exception as e:
            print(f"Erro ao importar produto: {e}")
            continue
    
    conn.commit()
    conn.close()
    
    # Limpar pasta temporaria
    if os.path.exists(TEMP_IMAGES_DIR):
        shutil.rmtree(TEMP_IMAGES_DIR)
    
    return imported_count


def ambiente_desktop_disponivel():
    """Indica se o servidor consegue abrir janelas desktop (GUI)."""
    if os.name == 'nt':
        return True
    return bool(os.environ.get('DISPLAY') or os.environ.get('WAYLAND_DISPLAY'))


def nome_ambiente_execucao():
    """Nome amigavel do ambiente atual para mensagens."""
    if os.environ.get('RENDER'):
        return 'Render'
    if os.environ.get('VERCEL'):
        return 'Vercel'
    if os.name == 'nt':
        return 'Windows'
    return 'Servidor Linux sem interface grafica'


def abrir_janela_sistema(script_name, nome_exibicao):
    """Abre um sistema desktop em uma nova janela sem bloquear o Flask."""
    script_path = os.path.join(BASE_DIR, script_name)
    module_name = os.path.splitext(script_name)[0]
    running_frozen = bool(getattr(sys, 'frozen', False))

    if not os.path.exists(script_path) and not running_frozen:
        raise FileNotFoundError(f"Arquivo nao encontrado: {script_name}")

    if not ambiente_desktop_disponivel():
        ambiente = nome_ambiente_execucao()
        raise RuntimeError(
            f"Nao e possivel abrir janela desktop no ambiente {ambiente}. Use a interface web."
        )

    popen_kwargs = {
        'cwd': BASE_DIR,
    }

    if os.name == 'nt':
        pythonw_path = os.path.join(os.path.dirname(sys.executable), 'pythonw.exe')
        # Usar python.exe em vez de pythonw.exe para ver logs no console
        executable = sys.executable
        # Nao usar CREATE_NEW_CONSOLE para manter logs no terminal atual
        # creationflags = getattr(subprocess, 'CREATE_NEW_CONSOLE', 0)
        # if creationflags:
        #     popen_kwargs['creationflags'] = creationflags
        # Redirecionar stdout/stderr para ver logs
        popen_kwargs['stdout'] = subprocess.PIPE
        popen_kwargs['stderr'] = subprocess.STDOUT
    else:
        executable = sys.executable
        popen_kwargs['start_new_session'] = True

    if os.path.exists(script_path):
        cmd = [executable, script_path]
    elif running_frozen:
        cmd = [executable, "--run-module", module_name]
    else:
        raise FileNotFoundError(f"Arquivo nao encontrado: {script_name}")

    print(f"[Flask] Abrindo: {' '.join(cmd)}")
    processo = subprocess.Popen(cmd, **popen_kwargs)
    
    # Thread para ler e imprimir logs do sistema
    def ler_logs():
        try:
            if processo.stdout:
                for linha in iter(processo.stdout.readline, b''):
                    if linha:
                        print(f"[{nome_exibicao}] {linha.decode('utf-8', errors='ignore').rstrip()}")
        except Exception as e:
            print(f"[Flask] Erro ao ler logs: {e}")
    
    threading.Thread(target=ler_logs, daemon=True).start()
    
    time.sleep(0.6)
    if processo.poll() is not None:
        raise RuntimeError(
            f"{nome_exibicao} encerrou logo apos iniciar (codigo {processo.returncode})."
        )

    return {
        'nome': nome_exibicao,
        'arquivo': script_name,
        'pid': processo.pid,
        'status': 'sucesso',
        'mensagem': f'{nome_exibicao} aberto em nova janela.'
    }

@app.route('/')
def index():
    """Pagina inicial atraente"""
    return render_template('index.html', config=SISTEMA_CONFIG)


@app.route("/sistema")
def sistema_dashboard():
    """Dashboard web (sem abrir janelas desktop)."""
    try:
        dados_sistema = _build_system_status()
        return render_template("sistema.html", config=SISTEMA_CONFIG, resultado=dados_sistema)
    except Exception as e:
        return render_template("iniciando.html", erro=str(e))


@app.route('/download-exe', methods=['POST'])
def download_exe():
    """Baixa o executavel desktop mais recente publicado em releases/."""
    exe = get_latest_desktop_exe()
    if not exe:
        debug_dirs = ", ".join([d for d in DESKTOP_RELEASES_DIRS])
        return (
            "Arquivo .exe ainda nao foi publicado. "
            f"Adicione o instalador em releases/ ou dist/ para liberar o download. Pastas verificadas: {debug_dirs}",
            404
        )

    # Em ambiente serverless (ex.: Vercel), arquivos grandes nao devem ser
    # retornados pela funcao Python; o ideal e servir como arquivo estatico.
    if IS_VERCEL:
        return redirect(f"/releases/{exe['filename']}")

    return send_file(exe['path'], as_attachment=True, download_name=exe['filename'])


@app.route('/download-exe', methods=['GET'])
def download_exe_get():
    """Evita disparos automaticos de download via GET."""
    return (
        "Download disponivel apenas pelo botao oficial. "
        "Clique em 'Baixar instalador .exe'.",
        405
    )

@app.route('/executar-sistema-legado')
def executar_sistema_legado():
    """Executa funcoes dos modulos dentro do mesmo processo Flask."""
    try:
        dados_sistema = _build_system_status()

        return render_template(
            'sistema.html',
            config=SISTEMA_CONFIG,
            resultado=dados_sistema
        )
    except Exception as e:
        return render_template('iniciando.html', erro=str(e))


@app.route('/executar-sistema')
@_login_required

def executar_sistema():
    """Abre as janelas do sistema original e/ou PLUS a partir do Flask."""
    try:
        # Executa o sistema Original primeiro quando não há target especificado
        if not request.args.get("target"):
            alvo = 'original'
        else:
            alvo = request.args.get('target', 'both').lower()
        scripts_por_alvo = {
            'original': [('sistema.py', 'Sistema Original')],
            'plus': [('sistema_plus.py', 'Sistema Plus')],
            'menu': [('menu_principal.py', 'Menu Principal')],
            'both': [
                ('sistema.py', 'Sistema Original'),
                ('sistema_plus.py', 'Sistema Plus')
            ],
            'all': [
                ('menu_principal.py', 'Menu Principal'),
                ('sistema.py', 'Sistema Original'),
                ('sistema_plus.py', 'Sistema Plus')
            ]
        }

        scripts_para_abrir = scripts_por_alvo.get(alvo, scripts_por_alvo['both'])
        ambiente_gui = ambiente_desktop_disponivel()
        ambiente_nome = nome_ambiente_execucao()
        dados_sistema = {
            'status': 'sucesso',
            'mensagem': 'As janelas do sistema foram abertas com sucesso.',
            'timestamp': datetime.now().isoformat(),
            'modulos_carregados': [],
            'janelas_abertas': [],
            'alvo': alvo,
            'ambiente_gui': ambiente_gui,
            'ambiente_nome': ambiente_nome,
            'build_info': BUILD_INFO,
            'distribuicao_desktop': get_desktop_distribution_info()
        }

        for script_name, nome_exibicao in scripts_para_abrir:
            try:
                janela = abrir_janela_sistema(script_name, nome_exibicao)
                dados_sistema['janelas_abertas'].append(janela)
                dados_sistema['modulos_carregados'].append(f'[OK] {script_name} aberto em nova janela')
            except Exception as e:
                dados_sistema['status'] = 'parcial'
                dados_sistema['janelas_abertas'].append({
                    'nome': nome_exibicao,
                    'arquivo': script_name,
                    'status': 'erro',
                    'mensagem': str(e)
                })
                dados_sistema['modulos_carregados'].append(f'[ERRO] {script_name}: {str(e)[:80]}')

        try:
            if sistema_plus is not None and hasattr(sistema_plus, 'contar_produtos_plus'):
                dados_sistema['total_produtos'] = sistema_plus.contar_produtos_plus()
            if sistema_plus is not None and hasattr(sistema_plus, 'contar_planilhas_plus'):
                dados_sistema['total_planilhas'] = sistema_plus.contar_planilhas_plus()

            if sistema_plus is not None:
                dados_sistema['modulos_carregados'].append('[OK] Estatisticas do sistema_plus.py carregadas')
            else:
                erro = ERROS_IMPORTACAO.get('sistema_plus', 'modulo nao carregado')
                dados_sistema['modulos_carregados'].append(
                    f'[ERRO] Estatisticas indisponiveis (sistema_plus.py): {erro[:80]}'
                )
        except Exception as e:
            dados_sistema['modulos_carregados'].append(f'[ERRO] Estatisticas do sistema_plus.py: {str(e)[:80]}')

        if not ambiente_gui:
            dados_sistema['status'] = 'erro'
            dados_sistema['mensagem'] = (
                f"Ambiente {ambiente_nome} nao suporta janelas desktop. "
                "Use apenas as rotas web."
            )

        if dados_sistema['janelas_abertas'] and all(
            item.get('status') == 'erro' for item in dados_sistema['janelas_abertas']
        ):
            dados_sistema['status'] = 'erro'
            if ambiente_gui:
                dados_sistema['mensagem'] = 'Nao foi possivel abrir as janelas solicitadas.'
        elif ambiente_gui and any(item.get('status') == 'erro' for item in dados_sistema['janelas_abertas']):
            dados_sistema['mensagem'] = 'Algumas janelas abriram e outras falharam.'

        return render_template('sistema.html',
                             config=SISTEMA_CONFIG,
                             resultado=dados_sistema)
    except Exception as e:
        return render_template('iniciando.html', erro=str(e))



@app.route('/teste-importacao-plus')
def teste_importacao_plus():
    """Rota de teste para verificar se a importação do PLUS funciona"""
    try:
        if sistema_plus is None:
            erro = ERROS_IMPORTACAO.get('sistema_plus', 'Modulo nao carregado')
            return jsonify({
                'status': 'erro',
                'mensagem': f'sistema_plus.py nao foi importado: {erro}'
            })
        
        # Testar se a função de importação existe
        if not hasattr(sistema_plus, 'importar_planilha_plus'):
            return jsonify({
                'status': 'erro',
                'mensagem': 'Funcao importar_planilha_plus nao existe'
            })
        
        # Verificar se o banco está configurado
        if hasattr(sistema_plus, 'get_cursor_plus'):
            try:
                cursor = sistema_plus.get_cursor_plus()
                cursor.execute("SELECT COUNT(*) FROM produtos_plus")
                total = cursor.fetchone()[0]
                return jsonify({
                    'status': 'ok',
                    'mensagem': 'sistema_plus.py carregado corretamente',
                    'total_produtos': total,
                    'funcao_importacao': True
                })
            except Exception as e:
                return jsonify({
                    'status': 'erro_banco',
                    'mensagem': f'Erro ao acessar banco: {str(e)}'
                })
        else:
            return jsonify({
                'status': 'erro',
                'mensagem': 'Funcao get_cursor_plus nao existe'
            })
            
    except Exception as e:
        return jsonify({
            'status': 'erro',
            'mensagem': f'Erro inesperado: {str(e)}'
        })


@app.route('/upload')
def upload_page():
    """Pagina de upload"""
    return render_template('upload.html', config=SISTEMA_CONFIG)

@app.route('/api/upload', methods=['POST'])
@_login_required
def upload_excel():
    """API para upload e processamento de Excel (isolado por organização)"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo invalido'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400
        
        # Salvar arquivo temporario
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(temp_path)
        
        # Processar Excel
        products_data, images = read_excel_with_images(temp_path)
        
        # Importar para banco vinculando à organização do usuário
        org_id = _user_org_id()
        imported_count = import_products_to_db(products_data, organization_id=org_id)
        
        # Limpar arquivo temporario
        os.remove(temp_path)
        
        return jsonify({
            'success': True,
            'message': f'Importacao concluida com sucesso!',
            'imported_count': imported_count,
            'images_found': len(images),
            'products_count': len(products_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/catalog")
@_login_required
def catalog():
    """Catálogo de produtos do banco PLUS - Isolado por empresa"""
    conn = get_db_connection()
    where_sql, params = _filtro_org_sql()
    sql = "SELECT * FROM produtos_plus"
    if where_sql:
        sql += f" WHERE {where_sql}"
    sql += " ORDER BY data_importacao DESC LIMIT 100"
    products = conn.execute(sql, params).fetchall()
    conn.close()
    
    return render_template("catalog.html", products=products, config=SISTEMA_CONFIG)

@app.route('/api/stats')
@_login_required
def get_stats():
    """API de estatisticas (isoladas por empresa)"""
    conn = get_db_connection()
    where_sql, params = _filtro_org_sql()
    where_clause = f" WHERE {where_sql}" if where_sql else ""
    where_and = (" AND " + where_sql) if where_sql else ""

    stats = {
        'total_products': conn.execute(
            f"SELECT COUNT(*) FROM produtos_plus{where_clause}", params
        ).fetchone()[0],
        'total_clients': conn.execute(
            f"SELECT COUNT(DISTINCT cliente) FROM produtos_plus{where_clause}", params
        ).fetchone()[0],
        'recent_imports': conn.execute(
            f"SELECT COUNT(*) FROM produtos_plus WHERE data_importacao >= date('now', '-7 days'){where_and}",
            params
        ).fetchone()[0]
    }
    
    conn.close()
    return jsonify(stats)


def _garantir_superadmin():
    """Garante que o superadmin exista no banco"""
    print("=== GARANTINDO SUPERADMIN ===")
    
    try:
        # Verificar se superadmin já existe
        test_user = _auth_user("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        if test_user:
            print("[OK] Superadmin já existe e funciona!")
            return
        
        # Criar superadmin
        print("[ERRO] Superadmin não encontrado, criando...")
        result = _ensure_superadmin("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        print(f"[OK] Superadmin criado com ID: {result}")
        
        # Testar novamente
        test_user = _auth_user("superadmin@planilhas.com", "GpA1XmI86lGB309W")
        if test_user:
            print("[OK] Superadmin criado e autenticado com sucesso!")
        else:
            print("[ERRO] ERRO: Superadmin criado mas não autentica!")
        
    except Exception as e:
        print(f"[ERRO] ERRO AO GARANTIR SUPERADMIN: {e}")
        import traceback
        traceback.print_exc()


def executar_modulo_desktop(module_name):
    """Executa um modulo desktop como se fosse script (__main__)."""
    try:
        runpy.run_module(module_name, run_name="__main__")
        return 0
    except Exception as e:
        print(f"[ERRO] Falha ao executar modulo {module_name}: {e}")
        return 1


def executar_app():
    """Inicializa o Flask em modo desktop/web sem abrir console de debug."""
    port = int(os.environ.get("PORT", 5000))
    debug_flag = False  # SEMPRE false
    # Modo rede local: PLANILHAS_LAN=1 permite acesso de outros PCs no mesmo WiFi
    lan_mode = os.environ.get('PLANILHAS_LAN', '0') == '1'
    if IS_VERCEL or IS_RENDER or lan_mode:
        host = '0.0.0.0'
    else:
        host = '127.0.0.1'
    running_desktop = (os.name == 'nt') and (not IS_VERCEL) and (not IS_RENDER)
    running_frozen = bool(getattr(sys, 'frozen', False))

    # No desktop, o executavel nao tem janela. Abrir o navegador evita "nao abriu nada".
    auto_open_browser = running_desktop and host == '127.0.0.1'

    log_desktop(f"Startup: frozen={running_frozen} cwd={os.getcwd()} base_dir={BASE_DIR} host={host} port={port}")
    
    # Mostrar URLs importantes no console
    print(f"\n{'='*60}")
    print(f">> SERVIDOR RODANDO!")
    print(f"{'='*60}")
    if host == '0.0.0.0':
        # Descobrir IP local para informar aos colegas de rede
        try:
            import socket as _sk
            _s = _sk.socket(_sk.AF_INET, _sk.SOCK_DGRAM)
            _s.connect(('8.8.8.8', 80))
            _local_ip = _s.getsockname()[0]
            _s.close()
        except Exception:
            _local_ip = '127.0.0.1'
        print(f"[LAN] Acesso em rede local: http://{_local_ip}:{port}/")
        print(f"[LOCAL] Acesso neste PC: http://127.0.0.1:{port}/")
    else:
        print(f"[IMG] Upload de Imagens: http://127.0.0.1:{port}/upload-imagens")
        print(f"[HOME] Página Inicial: http://127.0.0.1:{port}/")
    print(f"{'='*60}\n")

    if auto_open_browser:
        abrir_navegador_quando_pronto(f"http://127.0.0.1:{port}", '127.0.0.1', port)

    try:
        app.run(
            debug=debug_flag,
            host=host,
            port=port,
            use_reloader=False
        )
    except OSError as e:
        # Se a porta ja estiver em uso, provavelmente ja existe uma instancia rodando.
        log_desktop(f"[ERRO] app.run falhou: {repr(e)}")
        addr_in_use = False
        try:
            if getattr(e, 'winerror', None) == 10048:
                addr_in_use = True
            elif getattr(e, 'errno', None) in (98, 48):
                addr_in_use = True
            elif 'address already in use' in str(e).lower():
                addr_in_use = True
        except Exception:
            addr_in_use = False

        if running_desktop and host == '127.0.0.1':
            try:
                abrir_navegador_quando_pronto(f"http://127.0.0.1:{port}", host, port, timeout_segundos=2)
            except Exception:
                pass

        # Se a porta ja estava ocupada, abrir o navegador e encerrar e' suficiente.
        if addr_in_use:
            return

        raise


def copiar_exe_para_desktop():
    """Copia o executavel para o Desktop do usuario se estiver rodando como EXE empacotado."""
    try:
        running_frozen = bool(getattr(sys, 'frozen', False))
        if not running_frozen:
            return  # Nao esta rodando como EXE empacotado
        
        # Caminho do executavel atual
        if hasattr(sys, '_MEIPASS'):
            # PyInstaller onefile - o EXE esta em sys.executable
            exe_atual = sys.executable
        else:
            exe_atual = sys.executable
        
        # Caminho do Desktop
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        if not os.path.exists(desktop_path):
            return
        
        # Nome do arquivo
        exe_name = os.path.basename(exe_atual)
        if not exe_name.endswith('.exe'):
            exe_name = 'Planilhas.exe'
        
        destino = os.path.join(desktop_path, exe_name)
        
        # Se ja existe no Desktop e eh o mesmo arquivo, nao copiar
        if os.path.exists(destino):
            if os.path.samefile(exe_atual, destino):
                return  # Ja esta rodando do Desktop
            # Verificar se o arquivo do Desktop eh mais recente
            if os.path.getmtime(destino) >= os.path.getmtime(exe_atual):
                return  # Desktop tem versao igual ou mais recente
        
        # Copiar para o Desktop
        import shutil
        shutil.copy2(exe_atual, destino)
        print(f"[OK] EXE copiado para Desktop: {destino}")
        
    except Exception as e:
        print(f"DEBUG: Erro ao copiar para Desktop: {e}")


# ============================
# ROTAS DE UPLOAD DE IMAGENS (MOVIDAS PARA ANTES DO if __name__)
# ============================

@app.route('/upload-imagens')
def upload_imagens_page():
    """Pagina de upload de imagens para gerar links diretos."""
    return render_template('upload_imagens_link.html')


@app.route('/upload-imagem', methods=['POST'])
def upload_imagem():
    """Recebe upload de imagem e retorna URL publica."""
    try:
        if 'imagem' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhuma imagem enviada'}), 400
        
        file = request.files['imagem']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'}), 400
        
        # Validar extensao
        extensao = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        extensoes_permitidas = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}
        
        if extensao not in extensoes_permitidas:
            return jsonify({'success': False, 'message': 'Extensao nao permitida. Use: PNG, JPG, JPEG, GIF, WEBP, BMP'}), 400
        
        # Criar pasta uploads se nao existir
        pasta_uploads = os.path.join(BASE_DIR, 'uploads')
        if not os.path.exists(pasta_uploads):
            os.makedirs(pasta_uploads)
        
        # Gerar nome unico
        timestamp = int(time.time())
        nome_seguro = f"{timestamp}_{file.filename}"
        caminho_arquivo = os.path.join(pasta_uploads, nome_seguro)
        
        # Salvar arquivo
        file.save(caminho_arquivo)
        
        # Gerar URL publica
        url_imagem = f"/uploads/{nome_seguro}"
        
        # Usar URL base correta (Render ou localhost)
        if IS_RENDER:
            host_url = "https://planilhas-1.onrender.com"
        else:
            host_url = request.host_url.rstrip('/')
        url_completa = f"{host_url}{url_imagem}"
        
        # Retornar URL imediatamente (SEM ESPERAR BANCO)
        print(f"[OK] Imagem salva: {caminho_arquivo}")
        print(f"[LINK] URL: {url_completa}")
        
        # Salvar no banco em SEGUNDO PLANO (não bloqueia resposta)
        def salvar_no_banco_async():
            try:
                from web_access_db import salvar_imagem
                salvar_imagem(
                    filename=nome_seguro,
                    nome_original=file.filename,
                    url=url_completa,
                    tipo=extensao,
                    tamanho=os.path.getsize(caminho_arquivo)
                )
            except Exception as e:
                print(f"[AVISO] Erro ao salvar no banco (não crítico): {e}")
        
        import threading
        threading.Thread(target=salvar_no_banco_async, daemon=True).start()
        
        return jsonify({
            'success': True,
            'url': url_completa,
            'filename': nome_seguro
        })
        
    except Exception as e:
        print(f"[ERRO] Erro no upload: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# Servir arquivos da pasta uploads
@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    """Serve imagens da pasta uploads."""
    pasta_uploads = os.path.join(BASE_DIR, 'uploads')
    return send_from_directory(pasta_uploads, filename)


@app.route('/gerenciar-imagens')
def gerenciar_imagens():
    """Pagina para listar e buscar imagens arquivadas."""
    from web_access_db import listar_imagens, buscar_imagens
    
    termo = request.args.get('q', '')
    if termo:
        imagens = buscar_imagens(termo, limit=100)
    else:
        imagens = listar_imagens(limit=100)
    
    return render_template('gerenciar_imagens.html', imagens=imagens, termo=termo)


# ============================
# LICENÇA DESKTOP (anti-pirataria)
# ============================

@app.route('/painel/licenca-desktop')
@_login_required
def painel_licenca_desktop():
    """Painel admin: gerencia a única licença desktop da empresa."""
    user = _current_user()
    if not _is_admin_or_superadmin(user):
        return "Apenas administradores podem gerenciar licença desktop.", 403
    if _desktop_license is None:
        return "Sistema de licença indisponível.", 503

    org_id = _user_org_id()
    if not org_id and not _is_superadmin(user):
        return "Sua conta não está vinculada a uma empresa.", 400

    lic = None
    if org_id:
        lic = _desktop_license.get_license_by_org(org_id)
    return render_template('licenca_desktop.html', licenca=lic, user=user)


@app.route('/api/licenca/gerar', methods=['POST'])
@_login_required
def api_licenca_gerar():
    """Admin gera (ou recupera) o token único da empresa."""
    user = _current_user()
    if not _is_admin_or_superadmin(user):
        return jsonify({"ok": False, "erro": "Sem permissão"}), 403
    if _desktop_license is None:
        return jsonify({"ok": False, "erro": "Indisponível"}), 503

    org_id = _user_org_id()
    if not org_id:
        return jsonify({"ok": False, "erro": "Conta sem empresa"}), 400

    try:
        lic = _desktop_license.create_or_get_license(org_id)
        return jsonify({"ok": True, "token": lic["token"],
                        "ativada": bool(lic.get("hardware_id"))})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route('/api/licenca/baixar')
@_login_required
def api_licenca_baixar():
    """Baixa um licenca.txt com o token (deve ficar ao lado do Planilhas.exe)."""
    from flask import Response
    user = _current_user()
    if not _is_admin_or_superadmin(user):
        return "Sem permissão", 403
    if _desktop_license is None:
        return "Indisponível", 503

    org_id = _user_org_id()
    if not org_id:
        return "Conta sem empresa", 400

    lic = _desktop_license.create_or_get_license(org_id)
    server_url = request.host_url.rstrip('/')
    conteudo = (
        "# Licença Planilhas Desktop\n"
        "# NÃO EDITE NEM COMPARTILHE ESTE ARQUIVO!\n"
        "# Coloque-o na MESMA pasta do Planilhas.exe\n"
        f"TOKEN={lic['token']}\n"
        f"SERVER={server_url}\n"
    )
    return Response(
        conteudo,
        mimetype='text/plain',
        headers={'Content-Disposition': 'attachment; filename=licenca.txt'}
    )


@app.route('/api/licenca/reset', methods=['POST'])
@_login_required
def api_licenca_reset():
    """Admin libera ativação em outro PC (limpa hardware_id)."""
    user = _current_user()
    if not _is_admin_or_superadmin(user):
        return jsonify({"ok": False, "erro": "Sem permissão"}), 403
    if _desktop_license is None:
        return jsonify({"ok": False, "erro": "Indisponível"}), 503

    org_id = _user_org_id()
    if not org_id:
        return jsonify({"ok": False, "erro": "Conta sem empresa"}), 400

    _desktop_license.reset_license(org_id)
    return jsonify({"ok": True, "mensagem": "Licença liberada para novo PC"})


@app.route('/api/licenca/ativar', methods=['POST'])
def api_licenca_ativar():
    """Endpoint público chamado pelo .exe na primeira execução."""
    if _desktop_license is None:
        return jsonify({"ok": False, "motivo": "Servidor indisponível"}), 503
    data = request.get_json(silent=True) or {}
    token = data.get('token')
    hardware_id = data.get('hardware_id')
    resultado = _desktop_license.activate_license(token, hardware_id)
    return jsonify(resultado)


@app.route('/api/licenca/verificar', methods=['POST'])
def api_licenca_verificar():
    """Endpoint público chamado pelo .exe a cada inicialização."""
    if _desktop_license is None:
        return jsonify({"ok": False, "motivo": "Servidor indisponível"}), 503
    data = request.get_json(silent=True) or {}
    token = data.get('token')
    hardware_id = data.get('hardware_id')
    resultado = _desktop_license.verify_license(token, hardware_id)
    return jsonify(resultado)


# ============================
# INICIALIZACAO
# ============================

if __name__ == '__main__':
    try:
        print("=== INICIALIZANDO APP PRINCIPAL ===")
        print("IS_RENDER:", IS_RENDER)
        
        # Copiar EXE para Desktop se estiver rodando como executavel empacotado
        copiar_exe_para_desktop()
        
        # Inicializar banco de acesso no Render
        if IS_RENDER:
            print("=== INICIALIZANDO BANCO NO RENDER ===")
            try:
                _init_access_db()
                print("[OK] Banco inicializado com sucesso")
            except Exception as e:
                print(f"[ERRO] ERRO AO INICIAR BANCO: {e}")
                import traceback
                traceback.print_exc()
                raise
            
            print("Garantindo superadmin...")
            try:
                _garantir_superadmin()
                print("[OK] Superadmin garantido")
            except Exception as e:
                print(f"[ERRO] ERRO AO GARANTIR SUPERADMIN: {e}")
                import traceback
                traceback.print_exc()
                raise
        
        if len(sys.argv) >= 3 and sys.argv[1] == "--run-module":
            sys.exit(executar_modulo_desktop(sys.argv[2]))
        
        # INICIAR O SERVIDOR FLASK
        print(">> Iniciando servidor Flask...")
        executar_app()
            
    except Exception as e:
        print(f"[ERRO] ERRO FATAL AO INICIAR APLICACAO: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


