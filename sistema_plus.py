import sqlite3
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText
import threading
from datetime import datetime
import csv
import json
from PIL import Image, ImageTk
import ctypes

# Base do projeto/arquivo para evitar variacao por CWD de .bat/.exe
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

from planilhas_paths import runtime_dir, ensure_from_resource, is_frozen, log_desktop

DATA_DIR = runtime_dir()
DB_PATH_PLUS = ensure_from_resource("banco_plus.db") if is_frozen() else os.path.join(BASE_DIR, "banco_plus.db")

def sanitizar_nome_arquivo(nome):
    """Gera um nome de arquivo seguro no Windows."""
    import re
    if nome is None:
        return ''
    nome = str(nome).strip()
    nome = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', nome)
    nome = re.sub(r'\\s+', ' ', nome).strip()
    nome = nome.strip(' ._')
    if len(nome) > 120:
        nome = nome[:120].rstrip(' ._')
    return nome

# ==============================
# BANCO DE DADOS - VERSÃƒO PLUS
# ==============================

# Thread-local storage para conexÃµes thread-safe
conn_local_plus = threading.local()
cursor_local_plus = threading.local()

def get_connection_plus():
    """ObtÃ©m conexÃ£o thread-safe para banco PLUS"""
    if not hasattr(conn_local_plus, 'conn'):
        conn_local_plus.conn = sqlite3.connect(DB_PATH_PLUS)
        conn_local_plus.conn.execute("PRAGMA journal_mode=DELETE")
        conn_local_plus.conn.execute("PRAGMA synchronous=FULL")
        conn_local_plus.conn.execute("PRAGMA cache_size=25000")
        conn_local_plus.conn.execute("PRAGMA temp_store=FILE")
        criar_banco_plus()
    
    return conn_local_plus.conn

def get_cursor_plus():
    """ObtÃ©m cursor thread-safe para banco PLUS"""
    if not hasattr(cursor_local_plus, 'cursor'):
        cursor_local_plus.cursor = get_connection_plus().cursor()
    return cursor_local_plus.cursor

def criar_banco_plus():
    """Cria o banco de dados PLUS com todas as tabelas e Ã­ndices"""
    cursor = get_cursor_plus()
    
    # Tabela principal com TODAS as 40 colunas e Ã­ndices otimizados
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS produtos_plus(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente TEXT COLLATE NOCASE,
        arquivo_origem TEXT,
        picture TEXT,
        imagem TEXT,
        link TEXT,
        codigo TEXT COLLATE NOCASE,
        descricao TEXT COLLATE NOCASE,
        peso TEXT,
        valor TEXT,
        ncm TEXT COLLATE NOCASE,
        doc TEXT,
        rev TEXT,
        code TEXT,
        quantity TEXT,
        um TEXT,
        ccy TEXT,
        total_amount TEXT,
        marca TEXT,
        inner_qty TEXT,
        master_qty TEXT,
        total_ctns TEXT,
        gross_weight TEXT,
        net_weight_pc TEXT,
        gross_weight_pc TEXT,
        net_weight_ctn TEXT,
        gross_weight_ctn TEXT,
        factory TEXT,
        address TEXT,
        telephone TEXT,
        ean13 TEXT,
        dun14_inner TEXT,
        dun14_master TEXT,
        length TEXT,
        width TEXT,
        height TEXT,
        cbm TEXT,
        prc_kg TEXT,
        li TEXT,
        obs TEXT,
        status TEXT,
        data_importacao TEXT,
        hash_dados TEXT
    )
    """)
    
    # Migracao automatica para bancos antigos sem as colunas fisicas
    cursor.execute("PRAGMA table_info(produtos_plus)")
    colunas_existentes = {row[1] for row in cursor.fetchall()}
    if 'picture' not in colunas_existentes:
        cursor.execute("ALTER TABLE produtos_plus ADD COLUMN picture TEXT DEFAULT ''")
        print("DEBUG: Coluna fisica 'picture' adicionada em produtos_plus")
    if 'imagem' not in colunas_existentes:
        cursor.execute("ALTER TABLE produtos_plus ADD COLUMN imagem TEXT DEFAULT ''")
        print("DEBUG: Coluna fisica 'imagem' adicionada em produtos_plus")
    if 'link' not in colunas_existentes:
        cursor.execute("ALTER TABLE produtos_plus ADD COLUMN link TEXT DEFAULT ''")
        print("DEBUG: Coluna fisica 'link' adicionada em produtos_plus")
    
    # Ãndices compostos para busca ultra-rÃ¡pida
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_cliente_plus ON produtos_plus(cliente)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_codigo_plus ON produtos_plus(codigo)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_descricao_plus ON produtos_plus(descricao)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ncm_plus ON produtos_plus(ncm)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_hash_plus ON produtos_plus(hash_dados)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_busca_completa_plus ON produtos_plus(cliente, codigo, descricao, ncm)")
    
    # Tabela de estatÃ­sticas
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS stats_plus(
        id INTEGER PRIMARY KEY,
        total_planilhas INTEGER,
        total_produtos INTEGER,
        ultima_atualizacao TEXT,
        tamanho_banco_mb REAL
    )
    """)
    
    get_connection_plus().commit()

# ==============================
# FUNÃ‡Ã•ES OTIMIZADAS
# ==============================

def formatar_valor_celula(cell, nome_coluna=''):
    """Formata o valor da célula preservando números e preços corretamente"""
    if cell is None:
        return ''
    
    # DEBUG: Mostrar o que está recebendo para colunas de valor
    coluna_lower = nome_coluna.lower() if nome_coluna else ''
    is_valor_col = any(palavra in coluna_lower for palavra in ['price', 'amount', 'valor', 'preco', 'preço', 'total', 'unit'])
    if is_valor_col and cell:
        print(f"DEBUG formatar_valor: coluna={nome_coluna}, tipo={type(cell)}, valor={repr(cell)}")
    
    # Converter para string para verificar se é fórmula
    cell_str = str(cell).strip() if cell else ''
    
    # Se for fórmula do Excel (começa com =), tentar extrair valor numérico
    if cell_str.startswith('='):
        # Tentar calcular fórmulas simples (ex: =40*6 ou =A1*B1)
        try:
            # Remover o = e avaliar expressão matemática simples
            formula = cell_str[1:]  # Remove o =
            # Para fórmulas simples como 40*6 ou 100+50
            import re
            # Remover letras (referências de células) e manter números e operadores
            numeros = re.findall(r'\d+\.?\d*', formula)
            if len(numeros) >= 2:
                # Tentar multiplicação simples (caso mais comum: =F40*40)
                resultado = 1
                for n in numeros:
                    resultado *= float(n)
                # Verificar se é coluna de valor/preco
                if is_valor_col:
                    resultado_str = f"{resultado:.2f}"
                    print(f"DEBUG formatar_valor: FÓRMULA CALCULADA {cell_str} = {resultado_str}")
                    return resultado_str
                return str(int(resultado)) if resultado == int(resultado) else str(resultado)
        except Exception as e:
            print(f"DEBUG formatar_valor: Erro ao calcular fórmula {cell_str}: {e}")
            pass
        # Se não conseguir calcular, retornar vazio para fórmulas
        return ''
    
    # Se for número (int ou float)
    if isinstance(cell, (int, float)):
        # Verificar se é quantidade (CTNS/QTY) - não formatar como valor
        is_quantidade = any(palavra in coluna_lower for palavra in ['ctns', 'qty', 'quantity', 'cartons', 'caixas'])
        
        # Se for coluna de preço/valor/amount, formatar com 2 casas decimais (mas não para quantidades)
        if is_valor_col and not is_quantidade:
            resultado = f"{cell:.2f}"
            print(f"DEBUG formatar_valor: NÚMERO FORMATADO {cell} -> {resultado}")
            return resultado
        else:
            # Para outros números, converter para string sem perder precisão
            if isinstance(cell, float):
                # Remover .0 se for inteiro
                if cell == int(cell):
                    return str(int(cell))
                return str(cell)
            return str(cell)
    
    # Se for string que parece número (ex: "240,00" ou "240.00")
    if is_valor_col:
        try:
            # Tentar converter string numérica
            # Substituir vírgula por ponto para conversão
            cell_num = cell_str.replace(',', '.')
            # Remover pontos de milhar se houver (ex: "1.234,56" -> "1234.56")
            if '.' in cell_num:
                partes = cell_num.split('.')
                if len(partes) == 2 and len(partes[1]) <= 2:  # Provavelmente decimal
                    pass  # já está correto
                elif len(partes) > 2:  # Tem pontos de milhar
                    cell_num = ''.join(partes[:-1]) + '.' + partes[-1]
            valor = float(cell_num)
            resultado = f"{valor:.2f}"
            print(f"DEBUG formatar_valor: STRING CONVERTIDA {cell_str} -> {resultado}")
            return resultado
        except:
            pass
    
    # Se for string, apenas limpar
    return cell_str

def importar_planilha_plus(caminho_arquivo, cliente=None, progress_callback=None, duplicatas=False):
    """ImportaÃ§Ã£o ultra-rÃ¡pida com controle de duplicatas"""
    print(f"DEBUG: Importar planilha PLUS: {caminho_arquivo}")
    
    if not cliente:
        cliente = os.path.basename(caminho_arquivo).replace('.xlsx', '').replace('.xls', '')
    
    print(f"DEBUG: Cliente: {cliente}")
    
    try:
        from openpyxl import load_workbook
        import hashlib
        
        wb = load_workbook(caminho_arquivo, read_only=True)
        ws = wb.active
        
        # Detectar cabeÃ§alhos - usar a linha com mais cÃ©lulas preenchidas (linhas 2-20)
        max_celulas = 0
        cabecalhos = []
        linha_cabecalho = 1
        
        for row_num in range(2, min(21, ws.max_row + 1)):  # ComeÃ§ar da linha 2
            linha = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
            celulas_preenchidas = [c for c in linha if c and str(c).strip()]
            
            # Se esta linha tem mais cÃ©lulas preenchidas que a anterior, use-a
            if len(celulas_preenchidas) > max_celulas:
                max_celulas = len(celulas_preenchidas)
                cabecalhos = [str(c).strip() for c in linha if c and str(c).strip()]
                linha_cabecalho = row_num
                
        print(f"DEBUG: CabeÃ§alhos encontrados na linha {linha_cabecalho}: {len(cabecalhos)} colunas")
        print(f"DEBUG: Primeiros cabeÃ§alhos: {cabecalhos[:10]}")
        
        colunas_detectadas = detectar_colunas_excel_plus(cabecalhos)
        import sys
        sys.stdout.flush()
        print(f"DEBUG: Colunas detectadas: {colunas_detectadas}")
        sys.stdout.flush()
        
        # Verificar se descricao foi detectada
        if 'descricao' in colunas_detectadas.values():
            print("DEBUG: [OK] Coluna 'descricao' foi detectada!")
        else:
            print("DEBUG: [ERRO] Coluna 'descricao' NAO foi detectada!")
            print(f"DEBUG: Valores mapeados: {list(colunas_detectadas.values())}")
        
        # Verificar se peso foi detectado
        if 'peso' in colunas_detectadas.values():
            print("DEBUG: [OK] Coluna 'peso' foi detectada!")
        else:
            print("DEBUG: [ERRO] Coluna 'peso' NAO foi detectada!")
        
        total_importados = 0
        total_duplicatas = 0
        data_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Batch maior para performance
        batch_size = 5000
        dados_batch = []
        
        # ComeÃ§ar a ler dados apÃ³s a linha de cabeÃ§alho
        linha_inicial_dados = linha_cabecalho + 1
        
        print(f"DEBUG: Iniciando leitura de dados da linha {linha_inicial_dados}")
        print(f"DEBUG: Total de linhas na planilha: {ws.max_row}")
        print(f"DEBUG: Colunas detectadas para mapeamento: {colunas_detectadas}")
        
        linhas_lidas = 0
        linhas_vazias = 0
        for row in ws.iter_rows(min_row=linha_inicial_dados, values_only=True):
            linhas_lidas += 1
            
            # DEBUG: Mostrar cada linha lida (primeiras 3)
            if linhas_lidas <= 3:
                print(f"DEBUG: Linha {linhas_lidas} lida: {row[:5]}...")
            
            if all(cell is None or str(cell).strip() == '' for cell in row):
                linhas_vazias += 1
                if linhas_lidas <= 3:
                    print(f"DEBUG: Linha {linhas_lidas} está VAZIA - ignorando")
                continue
            
            # PULAR LINHA DE TOTAIS/RODAPÉ - não importar como produto
            # Detectar pela coluna CÓDIGO (índice 0 normalmente) ou se todas células são fórmulas de soma
            codigo_cell = str(row[0]) if len(row) > 0 else ''
            if codigo_cell.strip().lower() == 'totais' or codigo_cell.strip().lower() == 'total':
                print(f"DEBUG: Linha {linhas_lidas} é linha de TOTAIS (código='{codigo_cell}') - IGNORANDO")
                continue
            
            # DEBUG: Mostrar primeira linha de dados válida
            if total_importados == 0:
                print(f"DEBUG: === PRIMEIRA LINHA DE DADOS VÁLIDA ===")
                print(f"DEBUG: Row data completa: {row}")
                print(f"DEBUG: Tamanho da row: {len(row)}")
            
            # Extrair dados usando mapeamento direto (índice -> coluna banco)
            dados = {col: '' for col in COLUNAS_BANCO_PLUS}  # Inicializar todas colunas vazias
            
            for i, cell in enumerate(row):
                if i in colunas_detectadas:
                    coluna_banco = colunas_detectadas[i]
                    # Se for PICTURE, usar exatamente da planilha
                    if coluna_banco == 'picture':
                        picture_valor = str(cell).strip() if cell is not None else ''
                        dados[coluna_banco] = picture_valor  # ← SEM fallback, exato da planilha
                        print(f"DEBUG: PICTURE da planilha: '{picture_valor}'")
                    elif coluna_banco == 'imagem':
                        imagem_valor = str(cell).strip() if cell is not None else ''
                        dados[coluna_banco] = imagem_valor
                        print(f"DEBUG: IMAGEM da planilha: '{imagem_valor}'")
                    elif coluna_banco == 'link':
                        link_valor = str(cell).strip() if cell is not None else ''
                        dados[coluna_banco] = link_valor
                        print(f"DEBUG: LINK da planilha: '{link_valor}'")
                    else:
                        dados[coluna_banco] = formatar_valor_celula(cell, coluna_banco)
            
            # REMOVIDO: Fallback automático - agora usa exatamente da planilha
            # if not dados.get('picture') and len(row) > 0 and row[0]:
            #     dados['picture'] = str(row[0]).strip()
            
            # Calcular TOTAL AMOUNT UMO = QUANTITY × UNIT PRICE UMO (SEMPRE calcular se tiver dados)
            tem_qty = dados.get('quantity') and str(dados.get('quantity')).strip()
            tem_valor = dados.get('valor') and str(dados.get('valor')).strip()
            
            if tem_qty and tem_valor:
                try:
                    def parse_br(valor):
                        if not valor:
                            return 0.0
                        # Se já é número, retorna direto
                        if isinstance(valor, (int, float)):
                            return float(valor)
                        valor_str = str(valor).strip()
                        # Só converte se tiver vírgula (formato brasileiro)
                        if ',' in valor_str:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                        return float(valor_str)
                    
                    quantity = parse_br(dados['quantity'])
                    
                    # DEBUG após definir parse_br
                    if total_importados < 5:
                        print(f"[DEBUG PLUS] qty_bruto='{dados.get('quantity')}', qty_parse={quantity}, valor={tem_valor}")
                    unit_price = parse_br(dados['valor'])
                    total_amount = quantity * unit_price
                    
                    if total_amount == int(total_amount):
                        dados['total_amount'] = f"{int(total_amount)},00"
                    else:
                        dados['total_amount'] = f"{total_amount:.2f}".replace('.', ',')
                    
                    print(f"[OK] TOTAL AMOUNT calculado: {dados['quantity']} × {dados['valor']} = {dados['total_amount']}")
                except Exception as e:
                    print(f"[ERRO] Erro no calculo TOTAL AMOUNT: {e}")
            elif total_importados < 5:
                print(f"[PULOU] TOTAL AMOUNT - qty={tem_qty}, valor={tem_valor}")
            
            # Calcular TOTAL CTNS (ex: INNER_QTY / MASTER_QTY ou usar total_ctns direto)
            # Por enquanto mantém o valor da planilha se existir
            
            # Calcular TOTAL NET WEIGHT (kg) = NET_WEIGHT_PC × QUANTITY
            tem_nw = dados.get('net_weight_pc') and str(dados.get('net_weight_pc')).strip()
            
            if total_importados < 5:
                print(f"[DEBUG PLUS] net_weight_pc={tem_nw}, net_weight_antes={dados.get('net_weight')}")
            
            if tem_qty and tem_nw:
                try:
                    def parse_br(valor):
                        if not valor:
                            return 0.0
                        # Se já é número, retorna direto
                        if isinstance(valor, (int, float)):
                            return float(valor)
                        valor_str = str(valor).strip()
                        # Só converte se tiver vírgula (formato brasileiro)
                        if ',' in valor_str:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                        return float(valor_str)
                    
                    net_weight_pc = parse_br(dados['net_weight_pc'])
                    quantity = parse_br(dados['quantity'])
                    # Calcular TOTAL NET (já em kg)
                    total_net = net_weight_pc * quantity
                    
                    if total_net == int(total_net):
                        dados['peso'] = f"{int(total_net)},00"
                    else:
                        dados['peso'] = f"{total_net:.2f}".replace('.', ',')
                    
                    print(f"[OK] TOTAL NET WEIGHT calculado: {dados['net_weight_pc']} × {dados['quantity']} = {dados['peso']}")
                except Exception as e:
                    print(f"[ERRO] Erro no calculo TOTAL NET WEIGHT: {e}")
            elif total_importados < 5:
                print(f"[PULOU] TOTAL NET WEIGHT - qty={tem_qty}, nw={tem_nw}")
            
            # TOTAL GROSS WEIGHT: Usar valor da planilha se existir, senão calcular
            total_gross_excel = dados.get('gross_weight') and str(dados.get('gross_weight')).strip()
            tem_gw_pc = dados.get('gross_weight_pc') and str(dados.get('gross_weight_pc')).strip()
            
            if total_importados < 5:
                print(f"[DEBUG PLUS] gross_weight_excel='{total_gross_excel}', gross_weight_pc='{tem_gw_pc}'")
            
            # Se já tem valor total na planilha, usar ele (formatar corretamente)
            if total_gross_excel and total_gross_excel not in ['', '0', '0,00', '0.00']:
                try:
                    def parse_br(valor):
                        if not valor:
                            return 0.0
                        if isinstance(valor, (int, float)):
                            return float(valor)
                        valor_str = str(valor).strip()
                        if ',' in valor_str:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                        return float(valor_str)
                    
                    # Normalizar o formato para brasileiro
                    val_num = parse_br(total_gross_excel)
                    if val_num == int(val_num):
                        dados['gross_weight'] = f"{int(val_num)},00"
                    else:
                        dados['gross_weight'] = f"{val_num:.2f}".replace('.', ',')
                    
                    print(f"[OK] TOTAL GROSS WEIGHT importado da planilha: {dados['gross_weight']}")
                except Exception as e:
                    print(f"[ERRO] Erro ao formatar TOTAL GROSS WEIGHT: {e}")
            # Se não tem total mas tem per-piece, calcular
            elif tem_qty and tem_gw_pc:
                try:
                    def parse_br(valor):
                        if not valor:
                            return 0.0
                        if isinstance(valor, (int, float)):
                            return float(valor)
                        valor_str = str(valor).strip()
                        if ',' in valor_str:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                        return float(valor_str)
                    
                    gross_weight_pc = parse_br(dados['gross_weight_pc'])
                    quantity = parse_br(dados['quantity'])
                    total_gross = gross_weight_pc * quantity
                    
                    if total_gross == int(total_gross):
                        dados['gross_weight'] = f"{int(total_gross)},00"
                    else:
                        dados['gross_weight'] = f"{total_gross:.2f}".replace('.', ',')
                    
                    print(f"[OK] TOTAL GROSS WEIGHT calculado: {dados['gross_weight_pc']} × {dados['quantity']} = {dados['gross_weight']}")
                except Exception as e:
                    print(f"[ERRO] Erro no calculo TOTAL GROSS WEIGHT: {e}")
            elif total_importados < 5:
                print(f"[PULOU] TOTAL GROSS WEIGHT - sem valor na planilha e sem dados para calcular")
            
            # Calcular TOTAL CBM = (LENGTH × WIDTH × HEIGHT × TOTAL CTNS) / 1.000.000
            # Convertendo mm³ para m³ (dividir por 1.000.000.000) e multiplicando por CTNS
            # Ou se já estiver em metros: (m × m × m) × CTNS
            tem_length = dados.get('length') and str(dados.get('length')).strip()
            tem_width = dados.get('width') and str(dados.get('width')).strip()
            tem_height = dados.get('height') and str(dados.get('height')).strip()
            tem_ctns = dados.get('total_ctns') and str(dados.get('total_ctns')).strip()
            
            if total_importados < 5:
                print(f"[DEBUG PLUS] length={tem_length}, width={tem_width}, height={tem_height}, ctns={tem_ctns}, cbm_antes={dados.get('cbm')}")
            
            if tem_length and tem_width and tem_height and tem_ctns:
                try:
                    def parse_br(valor):
                        if not valor:
                            return 0.0
                        # Se já é número, retorna direto
                        if isinstance(valor, (int, float)):
                            return float(valor)
                        valor_str = str(valor).strip()
                        # Só converte se tiver vírgula (formato brasileiro)
                        if ',' in valor_str:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                        return float(valor_str)
                    
                    length = parse_br(dados['length'])
                    width = parse_br(dados['width'])
                    height = parse_br(dados['height'])
                    ctns = parse_br(dados['total_ctns'])
                    
                    # CBM = (L × W × H) / 1.000.000.000 (mm³ → m³) × CTNS
                    # Ou assumir que já está em metros: L × W × H × CTNS
                    cbm = (length * width * height * ctns) / 1000000000
                    
                    if cbm == int(cbm):
                        dados['cbm'] = f"{int(cbm)},00"
                    else:
                        dados['cbm'] = f"{cbm:.4f}".replace('.', ',')
                    
                    print(f"[OK] TOTAL CBM calculado: ({length} x {width} x {height} x {ctns}) / 1000000000 = {dados['cbm']}")
                except Exception as e:
                    print(f"[ERRO] Erro no calculo TOTAL CBM: {e}")
            elif total_importados < 5:
                print(f"[PULOU] TOTAL CBM - dados insuficientes")
            
            # Adicionar campos fixos
            dados['cliente'] = cliente
            dados['arquivo_origem'] = os.path.basename(caminho_arquivo)
            dados['data_importacao'] = data_atual
            
            # DEBUG: Mostrar primeiros dados sendo importados
            if total_importados < 3:
                print(f"DEBUG: Dados linha {total_importados + 1}: codigo='{dados.get('codigo', '')}', descricao='{dados.get('descricao', '')}'")
            
            # Criar hash para detectar duplicatas (usando codigo e descricao)
            hash_dados = hashlib.md5(
                f"{cliente}_{dados.get('codigo','')}_{dados.get('descricao','')}".encode()
            ).hexdigest()
            dados['hash_dados'] = hash_dados
            
            # Verificar duplicata se necessÃ¡rio
            if duplicatas:
                get_cursor_plus().execute("SELECT id FROM produtos_plus WHERE hash_dados = ?", (hash_dados,))
                if get_cursor_plus().fetchone():
                    total_duplicatas += 1
                    continue
            
            # Adicionar todas as colunas no batch (ordem do INSERT)
            dados_batch.append((
                dados['cliente'],
                dados['arquivo_origem'],
                dados['picture'],
                dados['imagem'],
                dados['link'],
                dados['codigo'],
                dados['descricao'],
                dados['peso'],
                dados['valor'],
                dados['ncm'],
                dados['doc'],
                dados['rev'],
                dados['code'],
                dados['quantity'],
                dados['um'],
                dados['ccy'],
                dados['total_amount'],
                dados['marca'],
                dados['inner_qty'],
                dados['master_qty'],
                dados['total_ctns'],
                dados['gross_weight'],
                dados['net_weight_pc'],
                dados['gross_weight_pc'],
                dados['net_weight_ctn'],
                dados['gross_weight_ctn'],
                dados['factory'],
                dados['address'],
                dados['telephone'],
                dados['ean13'],
                dados['dun14_inner'],
                dados['dun14_master'],
                dados['length'],
                dados['width'],
                dados['height'],
                dados['cbm'],
                dados['prc_kg'],
                dados['li'],
                dados['obs'],
                dados['status'],
                dados['data_importacao'],
                dados['hash_dados']
            ))
            
            # DEBUG: Mostrar quando primeiro item é adicionado
            if len(dados_batch) == 0:
                print(f"DEBUG: Primeiro item sendo adicionado ao batch - codigo: {dados.get('codigo', 'N/A')}")
            
            # Insert em batch com todas as colunas
            if len(dados_batch) >= batch_size:
                try:
                    get_cursor_plus().executemany("""
                        INSERT INTO produtos_plus
                        (cliente, arquivo_origem, picture, imagem, link, codigo, descricao, peso, valor, ncm, doc, rev, code,
                         quantity, um, ccy, total_amount, marca, inner_qty, master_qty,
                         total_ctns, gross_weight, net_weight_pc, gross_weight_pc,
                         net_weight_ctn, gross_weight_ctn, factory, address, telephone,
                         ean13, dun14_inner, dun14_master, length, width, height, cbm,
                         prc_kg, li, obs, status, data_importacao, hash_dados)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, dados_batch)
                    get_connection_plus().commit()
                    total_importados += len(dados_batch)
                    print(f"DEBUG: Batch inserido - Total: {total_importados}")
                except Exception as e_insert:
                    print(f"DEBUG: ❌ ERRO NO BATCH: {e_insert}")
                    print(f"DEBUG: Tamanho do batch: {len(dados_batch)}")
                    print(f"DEBUG: Primeiro item do batch: {dados_batch[0] if dados_batch else 'VAZIO'}")
                    raise
                dados_batch = []
                
                if progress_callback and total_importados % 10000 == 0:
                    progress_callback(f"Importados: {total_importados:,}")
        
        # Insert final com todas as colunas
        if dados_batch:
            try:
                get_cursor_plus().executemany("""
                    INSERT INTO produtos_plus
                    (cliente, arquivo_origem, picture, imagem, link, codigo, descricao, peso, valor, ncm, doc, rev, code,
                     quantity, um, ccy, total_amount, marca, inner_qty, master_qty,
                     total_ctns, gross_weight, net_weight_pc, gross_weight_pc,
                     net_weight_ctn, gross_weight_ctn, factory, address, telephone,
                     ean13, dun14_inner, dun14_master, length, width, height, cbm,
                     prc_kg, li, obs, status, data_importacao, hash_dados)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, dados_batch)
                get_connection_plus().commit()
                total_importados += len(dados_batch)
                print(f"DEBUG: Batch final inserido - Total: {total_importados}")
            except Exception as e_insert:
                print(f"DEBUG: ❌ ERRO NO INSERT FINAL: {e_insert}")
                print(f"DEBUG: Tamanho do batch: {len(dados_batch)}")
                raise
        
        wb.close()
        
        # DEBUG: Mostrar totais
        print(f"DEBUG: === FIM DA IMPORTAÇÃO ===")
        print(f"DEBUG: Linhas lidas: {linhas_lidas}")
        print(f"DEBUG: Linhas vazias ignoradas: {linhas_vazias}")
        print(f"DEBUG: Linhas válidas processadas: {linhas_lidas - linhas_vazias}")
        print(f"DEBUG: Total importados: {total_importados}")
        print(f"DEBUG: Total duplicatas: {total_duplicatas}")
        print(f"DEBUG: Tamanho do último batch: {len(dados_batch)}")
        
        # Atualizar estatÃ­sticas
        atualizar_estatisticas_plus()
        
        # DEBUG: Verificar dados no banco
        print(f"DEBUG: Verificando dados no banco PLUS apÃ³s importaÃ§Ã£o...")
        get_cursor_plus().execute("SELECT COUNT(*) FROM produtos_plus")
        total_banco = get_cursor_plus().fetchone()[0]
        print(f"DEBUG: Total de registros no banco PLUS: {total_banco}")
        
        get_cursor_plus().execute("SELECT codigo, descricao, picture, imagem, link FROM produtos_plus LIMIT 3")
        resultados_verificacao = get_cursor_plus().fetchall()
        print(f"DEBUG: Primeiros 3 registros no banco:")
        for r in resultados_verificacao:
            print(f"  codigo={r[0]}, descricao={r[1][:30] if r[1] else ''}, picture={r[2]}, imagem={r[3]}, link={r[4]}")
        
        return total_importados, total_duplicatas
        
    except Exception as e:
        print(f"Erro ao importar {caminho_arquivo}: {str(e)}")
        return 0, 0

# Definir colunas fixas do sistema PLUS (mesma ordem do banco)
COLUNAS_BANCO_PLUS = [
    'cliente', 'arquivo_origem', 'picture', 'imagem', 'link', 'codigo', 'descricao', 'peso', 'valor', 'ncm',
    'doc', 'rev', 'code', 'quantity', 'um', 'ccy', 'total_amount', 'marca',
    'inner_qty', 'master_qty', 'total_ctns', 'gross_weight', 'net_weight_pc',
    'gross_weight_pc', 'net_weight_ctn', 'gross_weight_ctn', 'factory',
    'address', 'telephone', 'ean13', 'dun14_inner', 'dun14_master',
    'length', 'width', 'height', 'cbm', 'prc_kg', 'li', 'obs', 'status',
    'data_importacao', 'hash_dados'
]

# Mapeamento de sinÃ´nimos para colunas PLUS
MAPEAMENTO_SINONIMOS_PLUS = {
    'picture': ['picture', 'imagem', 'image', 'foto', 'url', 'image url', 'img'],
    'imagem': ['imagem', 'image', 'foto', 'picture file', 'arquivo imagem'],
    'link': ['link', 'url', 'link imagem', 'url imagem', 'imagem url', 'image link', 'foto link', 'link foto'],
    'codigo': ['codigo', 'cÃ³digo', 'cod', 'code', 'item', 'sku', 'referencia', 'referÃªncia', 'id', 'produto_id'],
    'descricao': [
        'descricao', 'descricao portugues', 'descricao portugues (description portuguese)', 'description portuguese',
        'descrição', 'descrição portugues', 'descrição portugues (description portuguese)',
        'descricao português', 'descricao português (description portuguese)',
        'descrição português', 'descrição português (description portuguese)',
        'description portuguese', 'descricao (description portuguese)',
        'produto', 'item_desc', 'name', 'descrição do produto', 'titulo', 'nome',
        'item description', 'product description', 'product name',
        'desc', 'descr', 'descricao_produto', 'product_desc'
    ],
    'peso': ['peso', 'weight', 'kg', 'quilos', 'peso_bruto', 'peso_liquido', 'massa', 'gr', 'total net weight( kg )', 'total net weight'],
    'valor': ['valor', 'preco', 'preÃ§o', 'price', 'unitario', 'unitÃ¡rio', 'custo', 'valor_unit', 'preco_unit', 'unit price umo'],
    'ncm': ['ncm', 'nomenclatura', 'codigo_ncm', 'ncm_sh', 'codigo_ncm_sh', 'nsh', 'codigo_nsh', 'hs code', 'hs code'],
    'doc': ['doc', 'documento'],
    'rev': ['rev', 'revisao', 'revisÃ£o'],
    'code': ['code', 'codigo_interno'],
    'quantity': ['quantity', 'quantidade', 'qtd'],
    'um': ['um', 'unidade', 'un'],
    'ccy': ['ccy', 'moeda', 'currency'],
    'total_amount': ['total amount', 'valor_total', 'total', 'total amount umo'],
    'marca': ['marca', 'brand', 'fabricante', 'marca (brand)'],
    'inner_qty': ['inner qty', 'quantidade_interna', 'qtd_interna', 'inner quantity'],
    'master_qty': ['master qty', 'quantidade_master', 'qtd_master', 'master quantity'],
    'total_ctns': ['total ctns', 'caixas', 'cartons', 'total ctns'],
    'gross_weight': ['gross weight', 'peso_bruto', 'total gross weight( kg )', 'total gross weight (kg)'],
    'net_weight_pc': ['net weight pc', 'peso_liquido_unit', 'net weight / pc( g )', 'net weight / pc (g)'],
    'gross_weight_pc': ['gross weight pc', 'peso_bruto_unit', 'gross weight / pc( g )', 'gross weight / pc (g)'],
    'net_weight_ctn': ['net weight ctn', 'peso_liquido_cx', 'net weight / ctn( kg )', 'net weight / ctn (kg)'],
    'gross_weight_ctn': ['gross weight ctn', 'peso_bruto_cx', 'gross weight / ctn( kg )', 'gross weight / ctn (kg)'],
    'factory': ['factory', 'fabrica', 'fÃ¡brica', 'name of factory'],
    'address': ['address', 'endereco', 'endereÃ§o', 'address of factory'],
    'telephone': ['telephone', 'telefone', 'tel', 'phone'],
    'ean13': ['ean13', 'ean', 'barcode'],
    'dun14_inner': ['dun-14 inner', 'dun14_interno', 'dun_inner'],
    'dun14_master': ['dun-14 master', 'dun14_master'],
    'length': ['length', 'comprimento', 'length ctn'],
    'width': ['width', 'largura', 'width ctn'],
    'height': ['height', 'altura', 'height ctn'],
    'cbm': ['cbm', 'metro_cubico', 'total cbm'],
    'prc_kg': ['prc/kg', 'preco_kg', 'prc/kg'],
    'li': ['li', 'licenca', 'licenÃ§a'],
    'obs': ['obs', 'observacoes', 'observaÃ§Ãµes', 'notas'],
    'status': ['status', 'situacao', 'situaÃ§Ã£o', 'status da compra']
}

def detectar_colunas_excel_plus(cabecalhos):
    """DetecÃ§Ã£o de colunas com mapeamento flexÃ­vel (sinÃ´nimos)"""
    mapeamento = {}
    colunas_nao_mapeadas = []
    
    for i, cab in enumerate(cabecalhos):
        cab_lower = cab.lower().strip()
        coluna_encontrada = None
        
        # Tentar mapeamento direto primeiro
        for col_banco in COLUNAS_BANCO_PLUS:
            if cab_lower == col_banco.lower():
                coluna_encontrada = col_banco
                break
        
        # Se nÃ£o encontrou, tentar por sinÃ´nimos
        if not coluna_encontrada:
            for col_banco, sinonimos in MAPEAMENTO_SINONIMOS_PLUS.items():
                if cab_lower in [s.lower() for s in sinonimos]:
                    coluna_encontrada = col_banco
                    break
        
        if coluna_encontrada:
            mapeamento[i] = coluna_encontrada
            print(f"[OK] Coluna {i}: '{cab}' -> MAPEADA como '{coluna_encontrada}'")
        else:
            colunas_nao_mapeadas.append((i, cab))
            print(f"[NAO] Coluna {i}: '{cab}' -> NAO MAPEADA")
    
    print(f"\n=== RESUMO DO MAPEAMENTO PLUS ===")
    print(f"[OK] Total mapeadas: {len(mapeamento)}")
    print(f"[INFO] Total nao mapeadas: {len(colunas_nao_mapeadas)}")
    if colunas_nao_mapeadas:
        print(f"[INFO] Colunas nao mapeadas: {colunas_nao_mapeadas}")
    print(f"================================")
    import sys
    sys.stdout.flush()
    
    return mapeamento

COLUNAS_PLUS_BUSCA = [
    'cliente', 'arquivo_origem', 'picture', 'imagem', 'link', 'data_importacao', 'codigo', 'descricao', 'peso',
    'valor', 'ncm', 'doc', 'rev', 'code', 'quantity', 'um', 'ccy', 'total_amount',
    'marca', 'inner_qty', 'master_qty', 'total_ctns', 'gross_weight',
    'net_weight_pc', 'gross_weight_pc', 'net_weight_ctn', 'gross_weight_ctn',
    'factory', 'address', 'telephone', 'ean13', 'dun14_inner', 'dun14_master',
    'length', 'width', 'height', 'cbm', 'prc_kg', 'li', 'obs', 'status'
]

# Exibicao no PLUS igual ao sistema.py (mesmos nomes e sequencia principais)
COLUNAS_PLUS_EXIBICAO = [
    ('cliente', 'CLIENTE'),
    ('arquivo_origem', 'ARQUIVO_ORIGEM'),
    ('data_importacao', 'DATA_IMPORTACAO'),
    ('picture', 'PICTURE'),
    ('imagem', 'IMAGEM'),
    ('link', 'LINK'),
    ('doc', 'DOC'),
    ('rev', 'REV'),
    ('codigo', 'ITEM'),
    ('code', 'CODE'),
    ('quantity', 'QUANTITY'),
    ('um', 'UM'),
    ('ccy', 'CCY'),
    ('valor', 'UNIT PRICE UMO'),
    ('total_amount', 'TOTAL AMOUNT UMO'),
    ('descricao', 'DESCRICAO PORTUGUES (DESCRIPTION PORTUGUESE)'),
    ('marca', 'MARCA (BRAND)'),
    ('inner_qty', 'INNER QUANTITY'),
    ('master_qty', 'MASTER QUANTITY'),
    ('total_ctns', 'TOTAL CTNS'),
    ('peso', 'TOTAL NET WEIGHT( KG )'),
    ('gross_weight', 'TOTAL GROSS WEIGHT( KG )'),
    ('net_weight_pc', 'NET WEIGHT / PC( G )'),
    ('gross_weight_pc', 'GROSS WEIGHT / PC( G )'),
    ('net_weight_ctn', 'NET WEIGHT / CTN( KG )'),
    ('gross_weight_ctn', 'GROSS WEIGHT / CTN( KG )'),
    ('factory', 'NAME OF FACTORY'),
    ('address', 'ADDRESS OF FACTORY'),
    ('telephone', 'TELEPHONE'),
    ('ean13', 'EAN13'),
    ('dun14_inner', 'DUN-14 INNER'),
    ('dun14_master', 'DUN-14 MASTER'),
    ('length', 'LENGTH CTN'),
    ('width', 'WIDTH CTN'),
    ('height', 'HEIGHT CTN'),
    ('cbm', 'TOTAL CBM'),
    ('ncm', 'HS CODE'),
    ('prc_kg', 'PRC/KG'),
    ('li', 'LI'),
    ('obs', 'OBS'),
    ('status', 'STATUS DA COMPRA')
]
COLUNAS_PLUS_UI = tuple([nome for _, nome in COLUNAS_PLUS_EXIBICAO])


def buscar_produtos_plus(termo='', cliente_filtro='Todos', planilha_filtro='Todas', limit=10000):
    """Busca produtos no PLUS em todas as colunas textuais."""
    print(f"DEBUG: Buscar PLUS - termo='{termo}' cliente='{cliente_filtro}' planilha='{planilha_filtro}'")
    sql_colunas = ", ".join([f'{origem} AS "{nome}"' for origem, nome in COLUNAS_PLUS_EXIBICAO])
    query = f"SELECT {sql_colunas} FROM produtos_plus WHERE 1=1"
    params = []

    if termo and termo != '*':
        conditions = " OR ".join([
            f"LOWER(COALESCE({col}, '')) LIKE LOWER(?)" for col in COLUNAS_PLUS_BUSCA
        ])
        query += f" AND ({conditions})"
        params.extend([f"%{termo}%"] * len(COLUNAS_PLUS_BUSCA))

    if cliente_filtro and cliente_filtro != "Todos":
        query += " AND LOWER(TRIM(COALESCE(cliente, ''))) = LOWER(TRIM(?))"
        params.append(cliente_filtro)

    if planilha_filtro and planilha_filtro != "Todas":
        query += " AND LOWER(TRIM(COALESCE(arquivo_origem, ''))) = LOWER(TRIM(?))"
        params.append(planilha_filtro)

    query += " ORDER BY cliente, arquivo_origem LIMIT ?"
    params.append(limit)

    try:
        print(f"DEBUG: Query PLUS: {query[:220]}...")
        print(f"DEBUG: Params PLUS: {params[:6]}{'...' if len(params) > 6 else ''}")
        get_cursor_plus().execute(query, params)
        resultados = get_cursor_plus().fetchall()
        print(f"DEBUG: Resultados PLUS: {len(resultados)}")
        return resultados
    except Exception as e:
        print(f"DEBUG: Erro na busca PLUS: {e}")
        return []


def listar_clientes_plus():
    """Lista clientes unicos do banco PLUS."""
    get_cursor_plus().execute("""
        SELECT DISTINCT TRIM(cliente)
        FROM produtos_plus
        WHERE TRIM(COALESCE(cliente, '')) <> ''
        ORDER BY TRIM(cliente)
    """)
    return [row[0] for row in get_cursor_plus().fetchall()]


def listar_planilhas_plus():
    """Lista planilhas (arquivo_origem) unicas do banco PLUS."""
    get_cursor_plus().execute("""
        SELECT DISTINCT TRIM(arquivo_origem)
        FROM produtos_plus
        WHERE TRIM(COALESCE(arquivo_origem, '')) <> ''
        ORDER BY TRIM(arquivo_origem)
    """)
    return [row[0] for row in get_cursor_plus().fetchall()]

def contar_produtos_plus():
    """Contagem otimizada"""
    get_cursor_plus().execute("SELECT COUNT(*) FROM produtos_plus")
    count = get_cursor_plus().fetchone()[0]
    print(f"DEBUG: Total de produtos PLUS: {count}")
    return count

def contar_planilhas_plus():
    """Contagem de planilhas Ãºnicas"""
    get_cursor_plus().execute("SELECT COUNT(DISTINCT arquivo_origem) FROM produtos_plus")
    count = get_cursor_plus().fetchone()[0]
    print(f"DEBUG: Total de planilhas PLUS: {count}")
    return count

def tamanho_banco_plus():
    """Tamanho do banco em MB"""
    get_cursor_plus().execute("SELECT page_count * page_size as size FROM pragma_page_count(), pragma_page_size()")
    tamanho_bytes = get_cursor_plus().fetchone()[0]
    return round(tamanho_bytes / (1024 * 1024), 2)

def atualizar_estatisticas_plus():
    """Atualiza tabela de estatÃ­sticas"""
    total_produtos = contar_produtos_plus()
    total_planilhas = contar_planilhas_plus()
    tamanho_mb = tamanho_banco_plus()
    
    get_cursor_plus().execute("""
        INSERT OR REPLACE INTO stats_plus (id, total_planilhas, total_produtos, ultima_atualizacao, tamanho_banco_mb)
        VALUES (1, ?, ?, ?, ?)
    """, (total_planilhas, total_produtos, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), tamanho_mb))
    get_connection_plus().commit()

def otimizar_banco_plus():
    """OtimizaÃ§Ã£o do banco para performance mÃ¡xima"""
    get_cursor_plus().execute("VACUUM")
    get_cursor_plus().execute("ANALYZE")
    get_connection_plus().commit()

# ==============================
# INTERFACE PLUS
# ==============================

class SistemaPlanilhasPlus:
    def __init__(self):
        # Configurar ID do aplicativo ANTES de criar a janela (para ícone na taskbar)
        try:
            if os.name == 'nt':  # Windows
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('planilhas.com.sistema.plus')
        except Exception as e:
            print(f"DEBUG: Erro ao configurar AppUserModelID: {e}")
        
        self.janela = tk.Tk()
        self.janela.title("Sistema Planilhas PLUS - Alta Capacidade")
        self.janela.geometry("1400x800")
        self.janela.configure(bg='#1a1a1a')
        
        # Ícone da janela
        try:
            icon_path = os.path.join(BASE_DIR, "icon.ico")
            if os.path.exists(icon_path):
                self.janela.iconbitmap(icon_path)
        except Exception as e:
            print(f"DEBUG: Erro ao configurar ícone: {e}")
        
        # VariÃ¡veis
        self.termo_busca = tk.StringVar()
        self.cliente_selecionado = tk.StringVar(value="Todos")
        self.planilha_selecionada = tk.StringVar(value="Todas")
        self.pasta_exportacoes = self.carregar_configuracao_exportacoes_plus()
        
        self.criar_interface_plus()
        self.atualizar_estatisticas_interface()
        self.atualizar_lista_clientes_plus()
        self.atualizar_lista_planilhas_plus()
        self.executar_busca_plus()
    
    def normalizar_pasta_exportacoes_plus(self, pasta):
        """Normaliza a pasta de exportacoes e evita exportacoes/exportacoes/..."""
        pasta_base = os.path.normpath(pasta) if pasta else 'exportacoes'
        drive, tail = os.path.splitdrive(pasta_base)
        is_abs = tail.startswith(os.sep)
        partes = [p for p in tail.split(os.sep) if p]
        
        partes_limpas = []
        for parte in partes:
            if (
                partes_limpas
                and partes_limpas[-1].lower() == 'exportacoes'
                and parte.lower() == 'exportacoes'
            ):
                continue
            partes_limpas.append(parte)
        
        if not partes_limpas:
            partes_limpas = ['exportacoes']
        elif partes_limpas[-1].lower() != 'exportacoes':
            partes_limpas.append('exportacoes')
        
        prefixo = ''
        if drive:
            prefixo = drive + (os.sep if is_abs else '')
        elif is_abs:
            prefixo = os.sep
        
        return prefixo + os.sep.join(partes_limpas)
    
    def carregar_configuracao_exportacoes_plus(self):
        """Carrega a configuracao da pasta exportacoes de um arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_exportacoes_plus.json")
        pasta_padrao = os.path.join(DATA_DIR, "exportacoes")
        
        if os.path.exists(arquivo_config):
            try:
                with open(arquivo_config, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    pasta = config.get('pasta_exportacoes', pasta_padrao)
                    pasta_norm = os.path.abspath(self.normalizar_pasta_exportacoes_plus(pasta))
                    os.makedirs(pasta_norm, exist_ok=True)
                    print(f"DEBUG CONFIG PLUS: arquivo_config={arquivo_config}")
                    print(f"DEBUG CONFIG PLUS: pasta_raw={pasta}")
                    print(f"DEBUG CONFIG PLUS: pasta_norm={pasta_norm}")
                    
                    # Corrige automaticamente configuracoes antigas quebradas
                    if pasta != pasta_norm:
                        self.salvar_configuracao_exportacoes_plus(pasta_norm)
                    
                    return pasta_norm
            except Exception as e:
                print(f"DEBUG: Erro ao carregar configuracao PLUS: {e}")
        
        os.makedirs(pasta_padrao, exist_ok=True)
        return pasta_padrao
    
    def salvar_configuracao_exportacoes_plus(self, pasta):
        """Salva a configuracao da pasta exportacoes em um arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_exportacoes_plus.json")
        try:
            pasta_norm = os.path.abspath(self.normalizar_pasta_exportacoes_plus(pasta))
            with open(arquivo_config, 'w', encoding='utf-8') as f:
                json.dump({'pasta_exportacoes': pasta_norm}, f, indent=4)
            print(f"DEBUG: Configuracao PLUS salva: {pasta_norm}")
            print(f"DEBUG CONFIG PLUS: arquivo_config={arquivo_config}")
        except Exception as e:
            print(f"DEBUG: Erro ao salvar configuracao PLUS: {e}")
    
    def escolher_pasta_exportacoes_plus(self):
        """Abre dialogo para escolher pasta de exportacoes"""
        pasta_escolhida = filedialog.askdirectory(
            title="Escolha onde criar a pasta de exportaÃ§Ãµes PLUS",
            initialdir=self.pasta_exportacoes
        )
        
        if pasta_escolhida:
            pasta_completa = os.path.abspath(self.normalizar_pasta_exportacoes_plus(pasta_escolhida))
            os.makedirs(pasta_completa, exist_ok=True)
            print(f"DEBUG CONFIG PLUS: pasta_escolhida={pasta_escolhida}")
            print(f"DEBUG CONFIG PLUS: pasta_final={pasta_completa}")
            
            self.pasta_exportacoes = pasta_completa
            self.salvar_configuracao_exportacoes_plus(pasta_completa)
            
            messagebox.showinfo("Sucesso", 
                f"Pasta de exportaÃ§Ãµes PLUS configurada em:\n{pasta_completa}\n\n"
                "Todas as exportaÃ§Ãµes serÃ£o salvas automaticamente aqui.")
    
    def criar_interface_plus(self):
        # Frame superior - Estatísticas PLUS (mais estreito)
        frame_stats_plus = tk.Frame(self.janela, bg='#2c3e50', height=90)
        frame_stats_plus.pack(fill='x', padx=10, pady=5)
        frame_stats_plus.pack_propagate(False)
        
        # Frame para logo e título (linha superior)
        frame_logo_titulo = tk.Frame(frame_stats_plus, bg='#2c3e50')
        frame_logo_titulo.pack(fill='x', pady=(5, 2))
        
        # Logo no canto esquerdo (menor)
        try:
            logo_path = os.path.join(BASE_DIR, "img", "Penacho laranja em fundo neutro.png")
            if os.path.exists(logo_path):
                # Carregar e redimensionar imagem (menor)
                logo_image = Image.open(logo_path)
                logo_image = logo_image.resize((50, 50), Image.Resampling.LANCZOS)
                logo_photo = ImageTk.PhotoImage(logo_image)
                
                logo_label = tk.Label(frame_logo_titulo, image=logo_photo, bg='#2c3e50')
                logo_label.image = logo_photo  # Manter referÃªncia
                logo_label.pack(side='left', padx=(15, 10))
            else:
                # Fallback emoji se imagem nao existir
                logo_label = tk.Label(frame_logo_titulo, text="🪶", font=('Arial', 36), bg='#2c3e50', fg='orange')
                logo_label.pack(side='left', padx=(15, 10))
        except Exception as e:
            # Fallback emoji em caso de erro
            logo_label = tk.Label(frame_logo_titulo, text="🪶", font=('Arial', 36), bg='#2c3e50', fg='orange')
            logo_label.pack(side='left', padx=(15, 10))
        
        # Título ao lado do logo (fonte menor)
        titulo_label = tk.Label(frame_logo_titulo, text="🚀 planilhas.com PLUS", font=('Arial', 18, 'bold'), bg='#2c3e50', fg='white')
        titulo_label.pack(side='left', padx=(0, 15))
        
        # Frame de navegação no canto direito
        frame_navegacao = tk.Frame(frame_logo_titulo, bg='#2c3e50')
        frame_navegacao.pack(side='right', padx=(20, 0))
        
        # Botão para voltar ao Sistema Original
        btn_voltar_original = tk.Button(frame_navegacao, text="📊 Sistema Original", 
                                       command=self.voltar_sistema_original,
                                       bg='#3498db', fg='white', font=('Arial', 10, 'bold'),
                                       relief='raised', bd=2, cursor='hand2')
        btn_voltar_original.pack(pady=2)
        
        # Estatísticas em linha separada (linha inferior) - mais espaço para não cortar
        frame_stats_container = tk.Frame(frame_stats_plus, bg='#2c3e50')
        frame_stats_container.pack(fill='x', pady=(5, 8))
        
        self.label_stats_plus = tk.Label(frame_stats_container, text="", bg='#2c3e50', fg='white', 
                                        font=('Arial', 11, 'bold'), height=2)
        self.label_stats_plus.pack(fill='x', padx=10)
        
        # Frame de controles avançados
        frame_controles = tk.LabelFrame(self.janela, text="⚙️ Controles Avançados", 
                                      font=('Arial', 14, 'bold'), bg='#34495e')
        frame_controles.pack(fill='x', padx=10, pady=5)
        
        # Linha 1 - Busca
        frame_busca = tk.Frame(frame_controles, bg='#34495e')
        frame_busca.pack(fill='x', padx=10, pady=10)
        
        tk.Label(frame_busca, text="🔍 Busca:", font=('Arial', 12), bg='#34495e', fg='white').pack(side='left', padx=5)
        self.campo_busca = tk.Entry(frame_busca, textvariable=self.termo_busca, font=('Arial', 12), width=50)
        self.campo_busca.pack(side='left', padx=5)
        
        tk.Button(frame_busca, text="🔎 Buscar", command=self.executar_busca_plus,
                 bg='#3498db', fg='white', font=('Arial', 12, 'bold')).pack(side='left', padx=5)
        
        tk.Button(frame_busca, text="🧹 Limpar", command=self.limpar_busca_plus,
                 bg='#e74c3c', fg='white', font=('Arial', 12), width=12, height=1).pack(side='left', padx=5)
        
        # Linha 2 - Importação
        frame_import = tk.Frame(frame_controles, bg='#34495e')
        frame_import.pack(fill='x', padx=10, pady=5)
        
        tk.Button(frame_import, text="📁 Importar Pasta", command=self.importar_pasta_plus,
                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'), width=20).pack(side='left', padx=5)
        
        tk.Button(frame_import, text="📂 Selecionar Arquivos", command=self.selecionar_arquivos_plus,
                 bg='#f39c12', fg='white', font=('Arial', 12, 'bold'), width=20).pack(side='left', padx=5)
        
        tk.Button(frame_import, text="⚡ Otimizar Banco", command=self.otimizar_banco_interface,
                 bg='#9b59b6', fg='white', font=('Arial', 12, 'bold'), width=20).pack(side='left', padx=5)
        
        tk.Button(frame_import, text="🔄 Limpar Banco", command=self.limpar_banco_plus,
                 bg='#e67e22', fg='white', font=('Arial', 12, 'bold'), width=18).pack(side='left', padx=5)
        
        # Filtros
        tk.Label(frame_import, text="Cliente:", font=('Arial', 12), bg='#34495e', fg='white').pack(side='left', padx=5)
        self.combo_clientes_plus = ttk.Combobox(frame_import, textvariable=self.cliente_selecionado, 
                                               state='readonly', width=30)
        self.combo_clientes_plus.pack(side='left', padx=5)
        self.combo_clientes_plus['values'] = ["Todos"]

        tk.Label(frame_import, text="Planilha:", font=('Arial', 12), bg='#34495e', fg='white').pack(side='left', padx=5)
        self.combo_planilhas_plus = ttk.Combobox(frame_import, textvariable=self.planilha_selecionada,
                                                 state='readonly', width=35)
        self.combo_planilhas_plus.pack(side='left', padx=5)
        self.combo_planilhas_plus['values'] = ["Todas"]
        
        # Frame de resultados
        frame_resultados = tk.LabelFrame(self.janela, text="📋 Resultados (Capacidade Máxima)", 
                                       font=('Arial', 14, 'bold'), bg='#34495e')
        frame_resultados.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Barra de exportação
        frame_export = tk.Frame(frame_resultados, bg='#34495e')
        frame_export.pack(fill='x', padx=5, pady=5)
        
        tk.Button(frame_export, text="📊 Exportar Excel", command=self.exportar_excel_plus,
                 bg='#8e44ad', fg='white', font=('Arial', 12), width=20, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_export, text="🗂️ Pasta Exportações", command=self.escolher_pasta_exportacoes_plus,
                 bg='#3498db', fg='white', font=('Arial', 12), width=20, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_export, text="📈 Estatísticas", command=self.mostrar_estatisticas_detalhadas,
                 bg='#2ecc71', fg='white', font=('Arial', 12)).pack(side='left', padx=5)
        
        self.label_resultados_plus = tk.Label(frame_export, text="0 resultados", bg='#34495e', 
                                            fg='white', font=('Arial', 12))
        self.label_resultados_plus.pack(side='right', padx=5)
        
        # CONTAINER DE COLUNAS - Compacto
        frame_colunas = tk.LabelFrame(frame_resultados, text="📋 Colunas Visíveis", 
                                      font=('Arial', 10, 'bold'), bg='#d5d8dc', fg='#2c3e50')
        frame_colunas.pack(fill='x', padx=5, pady=2)
        
        # Frame scrollável para os checkboxes (altura reduzida)
        canvas_colunas = tk.Canvas(frame_colunas, bg='#d5d8dc', height=50, highlightthickness=0)
        scroll_colunas = ttk.Scrollbar(frame_colunas, orient='horizontal', command=canvas_colunas.xview)
        frame_checks = tk.Frame(canvas_colunas, bg='#d5d8dc')
        
        canvas_colunas.configure(xscrollcommand=scroll_colunas.set)
        canvas_colunas.pack(fill='x', expand=True)
        scroll_colunas.pack(fill='x')
        
        canvas_colunas.create_window((0, 0), window=frame_checks, anchor='nw')
        
        # Atualizar scrollregion quando o conteúdo mudar
        def atualizar_scrollregion(event=None):
            canvas_colunas.configure(scrollregion=canvas_colunas.bbox('all'))
        
        frame_checks.bind('<Configure>', atualizar_scrollregion)
        
        # Permitir scroll com mouse
        def on_mousewheel(event):
            canvas_colunas.xview_scroll(int(-1*(event.delta/120)), 'units')
        
        canvas_colunas.bind('<MouseWheel>', on_mousewheel)
        frame_checks.bind('<MouseWheel>', on_mousewheel)
        
        # Tabela otimizada
        frame_tabela = tk.Frame(frame_resultados)
        frame_tabela.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.colunas_plus = COLUNAS_PLUS_UI
        self.tabela_plus = ttk.Treeview(frame_tabela, columns=self.colunas_plus, show='headings', height=20)
        
        for col in self.colunas_plus:
            self.tabela_plus.heading(col, text=col.replace('_', ' '))
            col_norm = col.lower()
            if 'descricao' in col_norm or 'description' in col_norm:
                self.tabela_plus.column(col, width=400)
            elif 'cliente' in col_norm:
                self.tabela_plus.column(col, width=150)
            elif 'arquivo' in col_norm:
                self.tabela_plus.column(col, width=250)
            else:
                self.tabela_plus.column(col, width=100)
        
        # Scrollbars da tabela
        scroll_v = ttk.Scrollbar(frame_tabela, orient='vertical', command=self.tabela_plus.yview)
        scroll_h = ttk.Scrollbar(frame_tabela, orient='horizontal', command=self.tabela_plus.xview)
        self.tabela_plus.configure(yscrollcommand=scroll_v.set, xscrollcommand=scroll_h.set)
        
        self.tabela_plus.grid(row=0, column=0, sticky='nsew')
        scroll_v.grid(row=0, column=1, sticky='ns')
        scroll_h.grid(row=1, column=0, sticky='ew')
        
        frame_tabela.grid_rowconfigure(0, weight=1)
        frame_tabela.grid_columnconfigure(0, weight=1)
        
        # Variáveis para checkboxes das colunas
        self.colunas_visiveis_plus = {}
        self.checks_colunas_plus = {}
        
        # Criar checkboxes para cada coluna
        colunas_disponiveis = list(self.colunas_plus)
        
        # Carregar configuração salva ou usar padrão
        colunas_salvas = self.carregar_configuracao_colunas_plus()
        if colunas_salvas:
            colunas_padrao = colunas_salvas
        else:
            colunas_padrao = ['CLIENTE', 'ITEM', 'DESCRICAO PORTUGUES (DESCRIPTION PORTUGUESE)', 'UNIT PRICE UMO', 'TOTAL AMOUNT UMO']
        
        for i, col in enumerate(colunas_disponiveis):
            var = tk.BooleanVar(value=col in colunas_padrao)
            self.colunas_visiveis_plus[col] = var
            
            chk = tk.Checkbutton(frame_checks, text=col.replace('_', ' '), variable=var,
                                bg='#d5d8dc', font=('Arial', 9), fg='#2c3e50',
                                selectcolor='#ffffff',
                                command=lambda c=col: self.atualizar_colunas_visiveis_plus())
            chk.pack(side='left', padx=5, pady=2)
            self.checks_colunas_plus[col] = chk
        
        # Botão para aplicar seleção
        btn_aplicar = tk.Button(frame_colunas, text="✓ Aplicar", 
                               command=self.atualizar_colunas_visiveis_plus,
                               bg='#27ae60', fg='white', font=('Arial', 8), width=8)
        btn_aplicar.pack(side='right', padx=5, pady=2)
        
        # Atualizar scrollregion
        frame_checks.update_idletasks()
        canvas_colunas.configure(scrollregion=canvas_colunas.bbox('all'))
        
        # Aplicar configuração inicial após criar a tabela
        self.janela.after(100, self.atualizar_colunas_visiveis_plus)
        
        # Barra de status
        self.status_bar_plus = tk.Label(self.janela, text="Sistema PLUS - Pronto", bd=1, 
                                       relief='sunken', anchor='w', bg='#2c3e50', fg='white')
        self.status_bar_plus.pack(side='bottom', fill='x')
    
    def atualizar_lista_clientes_plus(self):
        """Atualiza os clientes no filtro, igual ao sistema original."""
        clientes = ["Todos"] + listar_clientes_plus()
        self.combo_clientes_plus['values'] = clientes
        if self.cliente_selecionado.get() not in clientes:
            self.cliente_selecionado.set("Todos")

    def atualizar_lista_planilhas_plus(self):
        """Atualiza a lista de planilhas importadas no filtro."""
        planilhas = ["Todas"] + listar_planilhas_plus()
        self.combo_planilhas_plus['values'] = planilhas
        if self.planilha_selecionada.get() not in planilhas:
            self.planilha_selecionada.set("Todas")

    def executar_busca_plus(self):
        termo = self.termo_busca.get().strip()
        cliente = self.cliente_selecionado.get()
        planilha = self.planilha_selecionada.get()
        print(f"DEBUG: Executar busca PLUS - termo='{termo}' cliente='{cliente}' planilha='{planilha}'")

        # Limpar tabela
        for item in self.tabela_plus.get_children():
            self.tabela_plus.delete(item)

        # Mesmo comportamento do sistema original:
        # sem termo e filtros padrao => listar tudo
        if not termo and (cliente != "Todos" or planilha != "Todas"):
            termo = "*"
        if not termo and cliente == "Todos" and planilha == "Todas":
            termo = "*"

        try:
            resultados = buscar_produtos_plus(
                termo,
                cliente_filtro=cliente,
                planilha_filtro=planilha,
                limit=10000
            )
            
            colunas_list = list(self.colunas_plus)
            print(f"DEBUG: Colunas disponíveis: {colunas_list}")
            
            # Calcular total da coluna TOTAL CTNS
            total_ctns = 0
            idx_total_ctns = None
            if 'TOTAL CTNS' in colunas_list:
                idx_total_ctns = colunas_list.index('TOTAL CTNS')
                print(f"DEBUG: idx_total_ctns = {idx_total_ctns}")
            else:
                print(f"DEBUG: 'TOTAL CTNS' NÃO encontrado em colunas_list")
            
            # Calcular total da coluna TOTAL AMOUNT UMO
            total_amount = 0.0
            idx_total_amount = None
            if 'TOTAL AMOUNT UMO' in colunas_list:
                idx_total_amount = colunas_list.index('TOTAL AMOUNT UMO')
                print(f"DEBUG: idx_total_amount = {idx_total_amount}")
            else:
                print(f"DEBUG: 'TOTAL AMOUNT UMO' NÃO encontrado")
            
            # Calcular total da coluna TOTAL NET WEIGHT (kg)
            total_net = 0.0
            idx_total_net = None
            # Verificar variações do nome da coluna (maiúsculas/minúsculas)
            idx_total_net = None
            for col in colunas_list:
                if 'TOTAL NET WEIGHT' in col.upper() and 'KG' in col.upper():
                    idx_total_net = colunas_list.index(col)
                    print(f"DEBUG: idx_total_net = {idx_total_net} (coluna='{col}')")
                    break
            if idx_total_net is None:
                print(f"DEBUG: 'TOTAL NET WEIGHT( kg )' NÃO encontrado. Colunas: {colunas_list}")
            
            # Calcular total da coluna TOTAL GROSS WEIGHT (KG)
            total_gross = 0.0
            idx_gross = None
            # Verificar variações do nome da coluna (maiúsculas/minúsculas)
            idx_gross = None
            for col in colunas_list:
                if 'TOTAL GROSS WEIGHT' in col.upper() and 'KG' in col.upper():
                    idx_gross = colunas_list.index(col)
                    print(f"DEBUG: idx_gross = {idx_gross} (coluna='{col}')")
                    break
            if idx_gross is None:
                print(f"DEBUG: 'TOTAL GROSS WEIGHT( kg )' NÃO encontrado. Colunas: {colunas_list}")
            
            # Calcular total da coluna TOTAL CBM
            total_cbm = 0.0
            idx_cbm = None
            for col in colunas_list:
                if 'CBM' in col.upper() or 'TOTAL CBM' in col.upper():
                    idx_cbm = colunas_list.index(col)
                    print(f"DEBUG: idx_cbm = {idx_cbm} (coluna='{col}')")
                    break
            if idx_cbm is None:
                print(f"DEBUG: 'TOTAL CBM' NÃO encontrado. Colunas: {colunas_list}")
            
            # Calcular total da coluna QUANTITY
            total_quantity = 0.0
            idx_quantity = None
            for col in colunas_list:
                if col.upper() == 'QUANTITY':
                    idx_quantity = colunas_list.index(col)
                    print(f"DEBUG: idx_quantity = {idx_quantity} (coluna='{col}')")
                    break
            if idx_quantity is None:
                print(f"DEBUG: 'QUANTITY' NÃO encontrado. Colunas: {colunas_list}")
            
            for i, resultado in enumerate(resultados):
                self.tabela_plus.insert('', 'end', values=resultado)
                # DEBUG: Mostrar primeira linha para verificar índices
                if i == 0:
                    print(f"[PLUS] DEBUG PRIMEIRA LINHA: idx_ctns={idx_total_ctns}, idx_amount={idx_total_amount}, idx_net={idx_total_net}, idx_gross={idx_gross}")
                    print(f"[PLUS] DEBUG VALORES LINHA 0: CTNS='{resultado[idx_total_ctns] if idx_total_ctns else 'N/A'}', Amount='{resultado[idx_total_amount] if idx_total_amount else 'N/A'}', Net='{resultado[idx_total_net] if idx_total_net else 'N/A'}', Gross='{resultado[idx_gross] if idx_gross else 'N/A'}'")
                # Somar CTNS se a coluna existir
                if idx_total_ctns is not None and idx_total_ctns < len(resultado):
                    try:
                        val_raw = resultado[idx_total_ctns]
                        val = str(val_raw).replace('.', '').replace(',', '.')
                        if val and val.replace('.','').isdigit():
                            total_ctns += float(val)
                            if i < 3:
                                print(f"[PLUS] DEBUG SOMA CTNS linha {i}: raw='{val_raw}' -> parsed={val} -> total={total_ctns}")
                    except Exception as e:
                        if i < 3:
                            print(f"[PLUS] DEBUG ERRO CTNS linha {i}: {e}")
                # Somar TOTAL AMOUNT se a coluna existir
                if idx_total_amount is not None and idx_total_amount < len(resultado):
                    try:
                        val_raw = resultado[idx_total_amount]
                        val = str(val_raw).replace('.', '').replace(',', '.')
                        if val:
                            total_amount += float(val)
                            if i < 3:
                                print(f"[PLUS] DEBUG SOMA AMOUNT linha {i}: raw='{val_raw}' -> parsed={val} -> total={total_amount}")
                    except Exception as e:
                        if i < 3:
                            print(f"[PLUS] DEBUG ERRO AMOUNT linha {i}: {e}")
                # Somar TOTAL NET WEIGHT se a coluna existir
                if idx_total_net is not None and idx_total_net < len(resultado):
                    try:
                        val_raw = resultado[idx_total_net]
                        val = str(val_raw).replace('.', '').replace(',', '.')
                        if val:
                            total_net += float(val)
                            if i < 3:
                                print(f"[PLUS] DEBUG SOMA NET linha {i}: raw='{val_raw}' -> parsed={val} -> total={total_net}")
                    except Exception as e:
                        if i < 3:
                            print(f"[PLUS] DEBUG ERRO NET linha {i}: {e}")
                # Somar GROSS WEIGHT se a coluna existir
                if idx_gross is not None and idx_gross < len(resultado):
                    try:
                        val_raw = resultado[idx_gross]
                        val = str(val_raw).replace('.', '').replace(',', '.')
                        if val:
                            total_gross += float(val)
                            if i < 3:
                                print(f"[PLUS] DEBUG SOMA GROSS linha {i}: raw='{val_raw}' -> parsed={val} -> total={total_gross}")
                    except Exception as e:
                        if i < 3:
                            print(f"[PLUS] DEBUG ERRO GROSS linha {i}: {e}")
                # Somar TOTAL CBM se a coluna existir
                if idx_cbm is not None and idx_cbm < len(resultado):
                    try:
                        val_raw = resultado[idx_cbm]
                        val = str(val_raw).replace('.', '').replace(',', '.')
                        if val:
                            total_cbm += float(val)
                            if i < 3:
                                print(f"[PLUS] DEBUG SOMA CBM linha {i}: raw='{val_raw}' -> parsed={val} -> total={total_cbm}")
                    except Exception as e:
                        if i < 3:
                            print(f"[PLUS] DEBUG ERRO CBM linha {i}: {e}")
                # Somar QUANTITY se a coluna existir
                if idx_quantity is not None and idx_quantity < len(resultado):
                    try:
                        val_raw = resultado[idx_quantity]
                        val = str(val_raw).replace('.', '').replace(',', '.')
                        if val:
                            total_quantity += float(val)
                            if i < 3:
                                print(f"[PLUS] DEBUG SOMA QUANTITY linha {i}: raw='{val_raw}' -> parsed={val} -> total={total_quantity}")
                    except Exception as e:
                        if i < 3:
                            print(f"[PLUS] DEBUG ERRO QUANTITY linha {i}: {e}")
            
            # Mostrar resultados + totais (SEMPRE mostrar, mesmo quando zero)
            print(f"DEBUG TOTAIS: CTNS={total_ctns}, Amount={total_amount}, Net={total_net}, Gross={total_gross}, CBM={total_cbm}, Qty={total_quantity}")
            texto_resultado = f"{len(resultados):,} resultados"
            texto_resultado += f" | Total Qty: {int(total_quantity)}"
            texto_resultado += f" | Total CTNS: {int(total_ctns)}"
            texto_resultado += f" | Total Amount: {total_amount:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            texto_resultado += f" | Total Net: {total_net:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            texto_resultado += f" | Total Gross: {total_gross:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            texto_resultado += f" | Total CBM: {total_cbm:,.4f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            print(f"DEBUG: texto_resultado = '{texto_resultado}'")
            self.label_resultados_plus.config(text=texto_resultado)
            
            # ADICIONAR LINHA DE TOTAIS NA TABELA
            try:
                # Criar lista de valores vazios para a linha de totais
                totais_row = [''] * len(colunas_list)
                # Preencher a coluna ITEM com 'TOTAL'
                if 'ITEM' in colunas_list:
                    totais_row[colunas_list.index('ITEM')] = '★ TOTAL'
                # Preencher as colunas de totais
                if idx_total_ctns is not None:
                    totais_row[idx_total_ctns] = int(total_ctns)
                if idx_total_amount is not None:
                    totais_row[idx_total_amount] = f"{total_amount:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                if idx_total_net is not None:
                    totais_row[idx_total_net] = f"{total_net:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                if idx_gross is not None:
                    totais_row[idx_gross] = f"{total_gross:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                if idx_cbm is not None:
                    totais_row[idx_cbm] = f"{total_cbm:,.4f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                if idx_quantity is not None:
                    totais_row[idx_quantity] = int(total_quantity)
                # Inserir linha de totais com tag especial
                self.tabela_plus.insert('', 'end', values=totais_row, tags=('total',))
                # Configurar cor de fundo para a linha de totais
                self.tabela_plus.tag_configure('total', background='#FFD700', font=('Arial', 9, 'bold'))
                print(f"DEBUG: Linha de totais inserida na tabela")
            except Exception as e:
                print(f"DEBUG: Erro ao inserir linha de totais: {e}")
            self.status_bar_plus.config(
                text=f"Busca PLUS concluida: {len(resultados):,} resultados | Cliente: {cliente} | Planilha: {planilha}"
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro na busca: {str(e)}")
    
    def limpar_busca_plus(self):
        self.termo_busca.set("")
        self.cliente_selecionado.set("Todos")
        self.planilha_selecionada.set("Todas")
        
        for item in self.tabela_plus.get_children():
            self.tabela_plus.delete(item)
        
        self.label_resultados_plus.config(text="0 resultados encontrados")
        self.status_bar_plus.config(text="Busca limpa")
    
    def limpar_banco_plus(self):
        """Limpa todos os dados do banco PLUS"""
        if messagebox.askyesno("Confirmar Limpeza", "Tem certeza que deseja limpar todos os dados? Esta ação não pode ser desfeita!"):
            try:
                cursor = get_cursor_plus()
                cursor.execute("DELETE FROM produtos_plus")
                get_connection_plus().commit()
                
                # Limpar a tabela
                for item in self.tabela_plus.get_children():
                    self.tabela_plus.delete(item)
                
                self.atualizar_estatisticas_plus()
                self.atualizar_lista_clientes_plus()
                self.limpar_busca_plus()
                
                messagebox.showinfo("Banco Limpo", "Todos os dados foram removidos com sucesso!")
                self.status_bar_plus.config(text="Banco de dados limpo")
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao limpar banco: {str(e)}")
    
    def atualizar_colunas_visiveis_plus(self):
        """Atualiza quais colunas são visíveis na tabela baseado nos checkboxes"""
        # Obter colunas selecionadas
        colunas_selecionadas = [col for col, var in self.colunas_visiveis_plus.items() if var.get()]
        
        if not colunas_selecionadas:
            messagebox.showwarning("Aviso", "Selecione pelo menos uma coluna!")
            return
        
        # Ocultar/mostrar colunas na tabela
        for col in self.colunas_plus:
            if col in colunas_selecionadas:
                # Restaurar largura original
                col_norm = col.lower()
                if 'descricao' in col_norm or 'description' in col_norm:
                    self.tabela_plus.column(col, width=400)
                elif 'cliente' in col_norm:
                    self.tabela_plus.column(col, width=150)
                elif 'arquivo' in col_norm:
                    self.tabela_plus.column(col, width=250)
                else:
                    self.tabela_plus.column(col, width=100)
            else:
                self.tabela_plus.column(col, width=0, stretch=False, minwidth=0)
        
        # Salvar configuração
        self.salvar_configuracao_colunas_plus(colunas_selecionadas)
        
        self.status_bar_plus.config(text=f"Colunas visíveis: {len(colunas_selecionadas)} de {len(self.colunas_plus)}")
    
    def salvar_configuracao_colunas_plus(self, colunas_selecionadas):
        """Salva as colunas visíveis em arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_colunas_plus.json")
        try:
            with open(arquivo_config, 'w', encoding='utf-8') as f:
                json.dump({'colunas_visiveis': colunas_selecionadas}, f, indent=4)
        except Exception as e:
            print(f"DEBUG: Erro ao salvar configuração de colunas PLUS: {e}")
    
    def carregar_configuracao_colunas_plus(self):
        """Carrega as colunas visíveis do arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_colunas_plus.json")
        if os.path.exists(arquivo_config):
            try:
                with open(arquivo_config, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get('colunas_visiveis', None)
            except Exception as e:
                print(f"DEBUG: Erro ao carregar configuração de colunas PLUS: {e}")
        return None
    
    def importar_pasta_plus(self):
        def importar_thread():
            import traceback
            try:
                print("[GUI] Iniciando importacao da pasta...")
                self.status_bar_plus.config(text="ImportaÃ§Ã£o PLUS em andamento...")
                self.janela.update()
                
                pasta = "planilhas"
                if not os.path.exists(pasta):
                    os.makedirs(pasta)
                
                arquivos = [f for f in os.listdir(pasta) if f.endswith(('.xlsx', '.xls'))]
                print(f"[GUI] Arquivos encontrados: {arquivos}")
                
                if not arquivos:
                    print("[GUI] Nenhum arquivo encontrado!")
                    messagebox.showwarning("Aviso", "Nenhuma planilha encontrada na pasta 'planilhas'!")
                    return
                
                total_geral = 0
                
                for arquivo in arquivos:
                    caminho = os.path.join(pasta, arquivo)
                    print(f"[GUI] Importando: {caminho}")
                    try:
                        importados, duplicatas = importar_planilha_plus(
                            caminho, 
                            cliente=None,  # Será extraído do nome do arquivo
                            progress_callback=self.atualizar_progresso_plus
                        )
                        print(f"[GUI] Resultado: {importados} importados, {duplicatas} duplicatas")
                        total_geral += importados
                    except Exception as e_arquivo:
                        print(f"[GUI] ERRO no arquivo {arquivo}: {e_arquivo}")
                        traceback.print_exc()
                        raise
                
                print(f"[GUI] Total geral: {total_geral}")
                self.atualizar_estatisticas_interface()
                self.atualizar_lista_clientes_plus()
                self.atualizar_lista_planilhas_plus()
                self.executar_busca_plus()
                messagebox.showinfo("ImportaÃ§Ã£o PLUS", f"Importados {total_geral:,} produtos!")
                self.status_bar_plus.config(text=f"ImportaÃ§Ã£o concluÃ­da: {total_geral:,} produtos")
                
            except Exception as e:
                erro_msg = str(e)
                erro_tb = traceback.format_exc()
                print(f"[GUI] ERRO GERAL: {erro_msg}")
                print(f"[GUI] TRACEBACK: {erro_tb}")
                messagebox.showerror("Erro", f"Erro na importaÃ§Ã£o: {erro_msg}")
        
        threading.Thread(target=importar_thread, daemon=True).start()
    
    def selecionar_arquivos_plus(self):
        arquivos = filedialog.askopenfilenames(
            title="Selecione planilhas (mÃºltiplo)",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        
        if arquivos:
            def importar_thread():
                import traceback
                try:
                    print(f"[GUI] Arquivos selecionados: {arquivos}")
                    total_geral = 0
                    for caminho in arquivos:
                        print(f"[GUI] Importando arquivo: {caminho}")
                        try:
                            importados, _ = importar_planilha_plus(
                                caminho, 
                                cliente=None,
                                progress_callback=self.atualizar_progresso_plus
                            )
                            print(f"[GUI] Resultado: {importados} importados")
                            total_geral += importados
                        except Exception as e_arquivo:
                            print(f"[GUI] ERRO no arquivo {caminho}: {e_arquivo}")
                            traceback.print_exc()
                            raise
                    
                    print(f"[GUI] Total geral: {total_geral}")
                    self.atualizar_estatisticas_interface()
                    self.atualizar_lista_clientes_plus()
                    self.atualizar_lista_planilhas_plus()
                    self.executar_busca_plus()
                    messagebox.showinfo("ImportaÃ§Ã£o PLUS", f"Importados {total_geral:,} produtos!")
                    
                except Exception as e:
                    erro_msg = str(e)
                    erro_tb = traceback.format_exc()
                    print(f"[GUI] ERRO GERAL: {erro_msg}")
                    print(f"[GUI] TRACEBACK: {erro_tb}")
                    messagebox.showerror("Erro", f"Erro na importaÃ§Ã£o: {erro_msg}")
            
            threading.Thread(target=importar_thread, daemon=True).start()
    
    def otimizar_banco_interface(self):
        try:
            self.status_bar_plus.config(text="Otimizando banco...")
            self.janela.update()
            
            otimizar_banco_plus()
            self.atualizar_estatisticas_interface()
            
            messagebox.showinfo("OtimizaÃ§Ã£o", "Banco otimizado com sucesso!")
            self.status_bar_plus.config(text="Banco otimizado")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro: {str(e)}")
    
    def exportar_excel_plus(self):
        """Exporta resultados PLUS para Excel/CSV - apenas planilha do filtro atual"""
        resultados = []
        for item in self.tabela_plus.get_children():
            resultados.append(self.tabela_plus.item(item)['values'])
        
        if not resultados:
            messagebox.showwarning("Aviso", "Sem resultados para exportar!")
            return
        
        try:
            pasta_destino = os.path.abspath(self.normalizar_pasta_exportacoes_plus(self.pasta_exportacoes))
            os.makedirs(pasta_destino, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            colunas_completas = list(self.colunas_plus)
            
            # Encontrar índices das colunas CLIENTE e ARQUIVO_ORIGEM
            idx_cliente = -1
            idx_arquivo = -1
            for i, col in enumerate(colunas_completas):
                if col.upper() == 'CLIENTE':
                    idx_cliente = i
                if col.upper() == 'ARQUIVO_ORIGEM':
                    idx_arquivo = i
            
            # Fallback se não encontrou
            if idx_cliente < 0:
                idx_cliente = 0
            if idx_arquivo < 0:
                idx_arquivo = 1
            
            # Pegar a planilha atual do filtro (primeira linha dos resultados)
            primeira_linha = resultados[0]
            cliente_atual = primeira_linha[idx_cliente] if idx_cliente < len(primeira_linha) else ''
            planilha_atual = primeira_linha[idx_arquivo] if idx_arquivo < len(primeira_linha) else ''
            
            # Exportar apenas essa planilha
            base_planilha = os.path.splitext(str(planilha_atual or 'resultado'))[0]
            base_planilha = sanitizar_nome_arquivo(base_planilha) or 'resultado'
            base_cliente = sanitizar_nome_arquivo(cliente_atual) or 'cliente'
            nome_saida = f"{base_planilha}__{base_cliente}__{timestamp}.csv"
            caminho = os.path.join(pasta_destino, nome_saida)
            
            # Evitar sobrescrever
            if os.path.exists(caminho):
                n = 2
                while True:
                    alt = os.path.join(pasta_destino, f"{base_planilha}__{base_cliente}__{timestamp}_{n}.csv")
                    if not os.path.exists(alt):
                        caminho = alt
                        break
                    n += 1
            
            with open(caminho, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(colunas_completas)
                writer.writerows(resultados)
                f.flush()
                os.fsync(f.fileno())
            
            if not os.path.exists(caminho):
                raise FileNotFoundError(f"Arquivo nao foi criado: {caminho}")
            
            self.status_bar_plus.config(text=f"Exportado: {caminho}")
            messagebox.showinfo("Exportado", f"Planilha exportada:\n{caminho}\n\nColunas: {len(colunas_completas)} | Linhas: {len(resultados)}")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro de Exportacao", f"Nao foi possivel exportar:\n{str(e)}")
    
    def exportar_csv_plus(self):
        """Exporta resultados PLUS para CSV (mesma função do Excel)"""
        self.exportar_excel_plus()

    def mostrar_estatisticas_detalhadas(self):
        stats_text = f"""
📊 ESTATÍSTICAS DETALHADAS

🗄️ Banco de Dados:
• Tamanho: {tamanho_banco_plus():.2f} MB
• Produtos: {contar_produtos_plus():,}
• Planilhas: {contar_planilhas_plus():,}

⚡ Performance:
• Cache: 50MB
• Memory Map: 256MB
• Índices: 6 otimizados

🚀 Capacidade:
• Máximo teórico: Ilimitado
• Recomendado: 5M+ produtos
• Testado: 1M+ produtos
        """
        messagebox.showinfo("Estatísticas PLUS", stats_text)

    def atualizar_estatisticas_interface(self):
        total_produtos = contar_produtos_plus()
        total_planilhas = contar_planilhas_plus()
        tamanho_mb = tamanho_banco_plus()
        
        stats_text = f"📦 Produtos: {total_produtos:,}  |  📁 Planilhas: {total_planilhas:,}  |  💾 Banco: {tamanho_mb:.1f}MB  |  ⚡ PLUS MODE"
        self.label_stats_plus.config(text=stats_text)
    
    def atualizar_progresso_plus(self, mensagem):
        self.status_bar_plus.config(text=mensagem)
        self.janela.update()
    
    def voltar_sistema_original(self):
        """Volta para o Sistema Original"""
        if messagebox.askyesno("Voltar para Sistema Original", 
                              "Deseja fechar o Sistema PLUS e abrir o Sistema Original?"):
            try:
                import subprocess
                import sys
                
                # Fecha o sistema atual
                self.janela.destroy()
                
                # Executa o sistema original
                subprocess.Popen([sys.executable, "sistema.py"])
                
            except Exception as e:
                messagebox.showerror("Erro", 
                    "NÃ£o foi possÃ­vel abrir o Sistema Original. Verifique se o arquivo 'sistema.py' existe na pasta.")
    
    def executar(self):
        """Inicia a interface PLUS"""
        self.janela.mainloop()

# ==============================
# INICIALIZAÃ‡ÃƒO
# ==============================

if __name__ == "__main__":
    app_plus = SistemaPlanilhasPlus()
    app_plus.executar()
